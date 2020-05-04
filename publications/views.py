from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.sites.shortcuts import get_current_site
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import Coalesce
from django.forms import modelformset_factory, ModelChoiceField
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from ast import literal_eval
from collections import Counter
from itertools import chain
from random import shuffle
from .tokens import account_activation_token
from .forms import AssessmentForm, AttributeForm, AttributeOptionForm, CoordinatesForm, DataForm, DataUploadForm, DateForm, EAVExperimentForm, EAVOutcomeForm, EAVPopulationForm, EAVPublicationForm, ExperimentForm1, ExperimentForm2, ExperimentDesignForm, ExperimentPopulationForm, ExperimentPopulationOutcomeForm, FullTextAssessmentForm, InterventionForm, KappaForm, OutcomeForm, ProfileForm, PublicationForm, PublicationPopulationForm, PublicationPopulationOutcomeForm, SignUpForm, StudyForm, UserForm, UserSubjectForm, XCountryForm
from .models import Assessment, AssessmentStatus, Attribute, Coordinates, Country, Crop, Data, Date, Design, EAV, Experiment, ExperimentDesign, ExperimentPopulation, ExperimentPopulationOutcome, Intervention, Outcome, Publication, PublicationPopulation, PublicationPopulationOutcome, Study, Subject, User, UserSubject, XCountry
from .serializers import AttributeSerializer, CountrySerializer, DataSerializer, DesignSerializer, EAVSerializer, ExperimentSerializer, ExperimentDesignSerializer, ExperimentPopulationSerializer, ExperimentPopulationOutcomeSerializer, InterventionSerializer, OutcomeSerializer, PublicationSerializer, PublicationPopulationSerializer, PublicationPopulationOutcomeSerializer, SubjectSerializer, UserSerializer
from .decorators import group_required
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from tempfile import NamedTemporaryFile
from mptt.forms import TreeNodeChoiceField
from haystack.generic_views import SearchView
from haystack.forms import SearchForm
from haystack.query import SearchQuerySet
from rest_framework import viewsets
import reversion
import csv
import json


class AttributeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for user-defined attributes (for the EAV model)
    """
    queryset = Attribute.objects.all()
    serializer_class = AttributeSerializer

    def get_queryset(self):
        queryset = Attribute.objects.all()
        parent_pk = self.request.GET.get('parent', '')
        if parent_pk is not '':
            queryset = Attribute.objects.get(pk=parent_pk).get_children()
        return queryset


class CountryViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "countries"
    """
    queryset = Country.objects.all()
    serializer_class = CountrySerializer


class DesignViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "designs"
    """
    queryset = Design.objects.all()
    serializer_class = DesignSerializer


class EAVViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "EAVs" (entity-attribute-values: user-defined fields)
    """
    queryset = EAV.objects.all()
    serializer_class = EAVSerializer


class ExperimentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "experiments" (i.e. one "intervention" in one "publication")
    """
    queryset = Experiment.objects.all()
    serializer_class = ExperimentSerializer


class ExperimentDesignViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "experiment_designs" (i.e. one "design" element in one "experiment")
    """
    queryset = ExperimentDesign.objects.distinct()
    serializer_class = ExperimentDesignSerializer


class ExperimentPopulationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "experiment_populations" (i.e. one "population" in one "experiment")
    """
    queryset = ExperimentPopulation.objects.all()
    serializer_class = ExperimentPopulationSerializer


class ExperimentPopulationOutcomeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "experiment populations outcomes" (i.e. one "outcome" for one "population" in one "experiment")
    """
    queryset = ExperimentPopulationOutcome.objects.all()
    serializer_class = ExperimentPopulationOutcomeSerializer


class DataViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for data (e.g., treatment mean, control mean, etc.)
    """
    queryset = Data.objects.all()
    serializer_class = DataSerializer

    def get_queryset(self):
        queryset = Data.objects.all()
        subject_pk = self.request.GET.get('subject', '')
        if subject_pk is not '':
            subject = Subject.objects.get(pk=subject_pk)
            subjects = subject.get_descendants(include_self=True)
            queryset = queryset.filter(subject__in=subjects)
        intervention_pk = self.request.GET.get('intervention', '')
        if intervention_pk is not '':
            interventions = Intervention.objects.get(pk=intervention_pk).get_descendants(include_self=True)
            queryset = queryset.filter(experiment__intervention__in=interventions)
        outcome_pk = self.request.GET.get('outcome', '')
        if outcome_pk is not '':
            outcomes = Outcome.objects.get(pk=outcome_pk).get_descendants(include_self=True)
            queryset = queryset.filter(experiment_population_outcome__outcome__in=outcomes)
        publication_pk = self.request.GET.get('publication', '')
        if publication_pk is not '':
            publication = Publication.objects.get(pk=publication_pk)
            queryset = queryset.filter(experiment__publication=publication)
        user_pk = self.request.GET.get('user', '')
        if user_pk is not '':
            user = User.objects.get(pk=user_pk)
            queryset = queryset.filter(experiment__user=user)
        return queryset.prefetch_related(
            'publication',
            'experiment', 'experiment__intervention', 'experiment__EAV_experiment', 'experiment__EAV_experiment__attribute', 'experiment__EAV_experiment__value_as_factor', 'experiment__xcountry_experiment', 'experiment__xcountry_experiment__country', 'experiment__study_experiment', 'experiment__experimentdesign_set',
            'experiment_population', 'experiment_population__population', 'experiment_population__EAV_population', 'experiment_population__EAV_population__attribute', 'experiment_population__EAV_population__value_as_factor', 'experiment_population__xcountry_population', 'experiment_population__xcountry_population__country', 'experiment_population__study_population',
            'experiment_population_outcome', 'experiment_population_outcome__outcome', 'experiment_population_outcome__EAV_outcome', 'experiment_population_outcome__EAV_outcome__attribute', 'experiment_population_outcome__EAV_outcome__value_as_factor', 'experiment_population_outcome__xcountry_outcome', 'experiment_population_outcome__xcountry_outcome__country', 'experiment_population_outcome__study_outcome',
            'publication__xcountry_publication', 'publication__EAV_publication'
        )


class InterventionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "interventions"
    """
    queryset = Intervention.objects.all()
    serializer_class = InterventionSerializer


class OutcomeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "outcomes"
    """
    queryset = Outcome.objects.all()
    serializer_class = OutcomeSerializer


class PublicationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "publications"
    """
    queryset = Publication.objects.all()
    serializer_class = PublicationSerializer


class PublicationPopulationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "publication_populations" (i.e. one "population" in one "publication")
    """
    queryset = PublicationPopulation.objects.all()
    serializer_class = PublicationPopulationSerializer


class PublicationPopulationOutcomeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "publication_population_outcomes" (i.e. one "outcome" for one "population" in one "publication")
    """
    queryset = PublicationPopulationOutcome.objects.all()
    serializer_class = PublicationPopulationOutcomeSerializer


class SubjectViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "subjects"
    """
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer

    def get_queryset(self):
        queryset = Subject.objects.all()
        subject_pk = self.request.GET.get('subject', '')
        if subject_pk is not '':
            subject = Subject.objects.get(pk=subject_pk)
            queryset = queryset.filter(subject=subject)
        return queryset


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for "users"
    """
    queryset = User.objects.all()
    serializer_class = UserSerializer


class MySearchView(SearchView):
    template_name = 'search/search.html'

    def get_queryset(self, *args, **kwargs):
        subject = self.kwargs['subject']
        subject = Subject.objects.get(slug=subject)
        queryset = super(MySearchView, self).get_queryset()
        return queryset.filter(subject=subject)

    def get_context_data(self, *args, **kwargs):
        context = super(MySearchView, self).get_context_data(*args, **kwargs)
        # do something
        return context


def subject(request, subject):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    context = {
        'subject': subject
    }
    # Get data for the sidebar.
    if user.is_authenticated:
        if UserSubject.objects.filter(user=user, subject=subject).exists():
            if Publication.objects.filter(subject=subject).exists():
                status = get_status(user, subject)
                context.update(status)
    return render(request, 'publications/subject.html', context)


def publications(request, subject, state='all', users='all_users', download='none'):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    if (users=='this_user'):
        users = User.objects.filter(pk=user.pk)
    else:
        users = User.objects.filter(usersubject__subject=subject)
    # All publications
    if (state == 'all'):
        publications = Publication.objects.filter(subject=subject).order_by('title')
    # Publications that this user has assessed as relevant based on title/abstract
    elif (state == 'relevant'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                is_relevant=True
            )
        ).order_by('title')
    # Publications that this user has assessed as not relevant based on title/abstract
    elif (state == 'not_relevant'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                is_relevant=False
            )
        ).order_by('title')
    # Publications that this user has not yet assessed based on title/abstract
    elif (state == 'not_assessed'):
        publications = Publication.objects.distinct().filter(subject=subject).exclude(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users
            )
        ).order_by('title')
    # Publications that this user has assessed as relevant based on full text
    elif (state == 'relevant_full_texts'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                is_relevant=True,  # is_relevant based on title/abstract
                full_text_is_relevant=True
            )
        ).order_by('title')
    # Publications that this user has assessed as not relevant based on full text
    elif (state == 'not_relevant_full_texts'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                is_relevant=True,  # is_relevant based on title/abstract
                full_text_is_relevant=False
            )
        ).order_by('title')
    elif (state == 'cannot_find'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                cannot_find=True
            )
        ).order_by('title')
    elif (state == 'cannot_access'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                cannot_access=True
            )
        ).order_by('title')
    elif (state == 'language_barrier'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                language_barrier=True
            )
        ).order_by('title')
    elif (state == 'secondary_literature'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                secondary_literature=True
            )
        ).order_by('title')
    elif (state == 'other'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                other=True
            )
        ).order_by('title')
    elif (state == 'no_pico'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users
            ).filter(
                Q(no_population=True) |
                Q(no_intervention=True) |
                Q(no_comparator=True) |
                Q(no_outcome=True)
            )
        ).order_by('title')
    # Publications that this user has marked as completed
    elif (state == 'completed'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                is_completed=True
            )
        ).order_by('title')
    # Publications that this user has not marked as completed
    elif (state == 'not_completed'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                is_relevant=True,
                is_completed=False
            )
        ).order_by('title')
        if (users.count() > 1):
            # Exclude publications that have been marked as completed by at least one person (even if they have not been marked as completed by other people).
            publications = publications.exclude(
                assessment__in=Assessment.objects.filter(
                    subject=subject,
                    user__in=users,
                    is_completed=True
                )
            )
    # Publications that have not yet been assessed at full-text stage
    elif (state == 'not_assessed_full_texts'):
        publications = Publication.objects.distinct().filter(
            assessment__in=Assessment.objects.filter(
                subject=subject,
                user__in=users,
                is_relevant=True,  # is_relevant based on title/abstract
                full_text_is_relevant=None
            )
        ).order_by('title')
        if (users.count() > 1):
            # Exclude publications that have been assessed by at least one person (even if they have not been assessed by other people).
            publications = publications.exclude(
                assessment__in=Assessment.objects.filter(
                    subject=subject,
                    user__in=users
                ).filter(
                    Q(full_text_is_relevant=True) | Q(full_text_is_relevant=False)
                )
            )
    # If the request is to download the publications, rather than viewing them online
    if download == 'CSV':
        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="bibliography.csv"'
        # Write the CSV
        writer = csv.writer(response, quoting=csv.QUOTE_ALL)
        writer.writerow(['Authors', 'Year', 'Title', 'Journal', 'Volume', 'Issue', 'DOI', 'URL', 'Abstract'])
        slug = subject.slug
        current_site = get_current_site(request)
        domain = current_site.domain
        for publication in publications:
            try:
                separator = ', '
                authors = separator.join(publication.author_list)
            except:
                authors = publication.authors
            path = reverse('publication', args=(), kwargs={'subject': slug, 'publication_pk': publication.pk})
            url = "http://{domain}{path}".format(domain=domain, path=path)
            writer.writerow([authors, publication.year, publication.title, publication.journal, publication.volume, publication.issue, publication.doi, url, publication.abstract])
        return response
    # If the request is not to download the publications
    page = request.GET.get('page', 1)
    paginator = Paginator(publications, 10)
    try:
        publications = paginator.page(page)
    except PageNotAnInteger:
        publications = paginator.page(1)
    except EmptyPage:
        publications = paginator.page(paginator.num_pages)
    context = {
        'subject': subject,
        'publications': publications
    }
    if user.is_authenticated:
        status = get_status(user, subject)
        context.update(status)
    return render(request, 'publications/publications.html', context)


def home(request):
    return render(request, 'publications/home.html')


def about(request):
    return render(request, 'publications/about.html')


def methods(request):
    return render(request, 'publications/methods.html')


def notes(request):
    return render(request, 'publications/notes.html')


def contact(request):
    return render(request, 'publications/contact.html')


def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():

            """
            ''' Begin reCAPTCHA validation '''

            recaptcha_response = request.POST.get('g-recaptcha-response')
            url = 'https://www.google.com/recaptcha/api/siteverify'
            values = {
                'secret': config.GOOGLE_RECAPTCHA_SECRET_KEY,
                'response': recaptcha_response
            }
            data = urllib.parse.urlencode(values).encode()
            req =  urllib.request.Request(url, data=data)
            response = urllib.request.urlopen(req)
            result = json.loads(response.read().decode())

            ''' End reCAPTCHA validation '''
            """

#            if result['success']:
            user = form.save()
            user.refresh_from_db()  # Load the profile instance created by the signal.
            user.profile.institution = form.cleaned_data.get('institution')
            user.is_active = False
            user.save()
            current_site = get_current_site(request)
            subject = 'Metadataset'
            message = render_to_string('publications/confirm_email.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)).decode(),
                'token': account_activation_token.make_token(user),
            })
            user.email_user(subject, message)
            return redirect('email_sent')
    else:
        form = SignUpForm()
    return render(request, 'publications/signup.html', {'form': form})


def email_sent(request):
    return render(request, 'publications/email_sent.html')


def email_confirmed(request):
    return render(request, 'publications/email_confirmed.html')


def email_not_confirmed(request):
    return render(request, 'publications/email_not_confirmed.html')


def confirm_email(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    if user is not None and account_activation_token.check_token(user, token):
        user.is_active = True
        user.profile.email_is_confirmed = True
        user.save()
        login(request, user)
        return render(request, 'publications/email_confirmed.html')
    else:
        return render(request, 'publications/email_not_confirmed.html', {'uid': uid})


@login_required
@transaction.atomic
def profile(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            user.first_name = user_form.cleaned_data.get('first_name')
            user.last_name = user_form.cleaned_data.get('last_name')
            user.profile.institution = profile_form.cleaned_data.get('institution')
            user.save()
            return render(request, 'publications/profile_updated.html')
    else:
        user_form = UserForm(instance=request.user)
        profile_form = ProfileForm(instance=request.user.profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form
    }

    return render(request, 'publications/profile.html', context)


@login_required
def publication(request, subject, publication_pk):
    """
    On this page, the user chooses interventions for this publication.
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    intervention = subject.intervention  # The root intervention for this subject (each subject can have its own classification of interventions)
    publication_pk = int(publication_pk)
    # This publication
    publication = get_object_or_404(Publication, pk=publication_pk, subject=subject)
    # Get data for the sidebar.
    status = get_status(user, subject)
    item = status.get('item')  # item = AssessmentStatus instance for this user and subject
    assessment_order = literal_eval(item.assessment_order)
    completed_assessments = literal_eval(item.completed_assessments)

    # The next pk and previous pk in assessment_order, to be used for navigation
    try:
        previous_pk = assessment_order[assessment_order.index(publication_pk) - 1]
    except:
        previous_pk = assessment_order[0]
    try:
        next_pk = assessment_order[assessment_order.index(publication_pk) + 1]
    except:
        next_pk = assessment_order[0]

    # "next_assessment" is the next pk in assessment_order that has not yet been assessed (whereas "next_pk" may or may not have been assessed).
    next_assessment = item.next_assessment
    if next_assessment == publication_pk:  # Users can pass from this "next_assessment" to the next "next_assessment".
        next_assessment = get_next_assessment(publication_pk, next_pk, assessment_order, completed_assessments)
        item.next_assessment = next_assessment
        item.save()

    # Form for this user_subject (for Kappa analysis)
    # This form is used twice, at different places in the template, and with different logic if request.POST (see below). That is why we have two versions of this form, instead of using a formset.
    user_subject_form_1 = UserSubjectForm(data=data, instance=user_subject, prefix="user_subject_form_1")
    user_subject_form_2 = UserSubjectForm(data=data, instance=user_subject, prefix="user_subject_form_2")
    # User_for_comparison choices for the form (users that have permission to work on this subject)
    user_subjects = UserSubject.objects.filter(subject=subject)
    user_subject_form_1.fields['user_for_comparison'] = ModelChoiceField(queryset=User.objects.filter(usersubject__in=user_subjects))
    user_subject_form_1.fields['user_for_comparison'].required = False
    user_subject_form_2.fields['user_for_comparison'] = ModelChoiceField(queryset=User.objects.filter(usersubject__in=user_subjects))
    user_subject_form_2.fields['user_for_comparison'].required = False
    # Form for this assessment
    if Assessment.objects.filter(publication=publication, user=user, subject=subject).exists():
        assessment = Assessment.objects.get(publication=publication, user=user, subject=subject)
        assessment_form = AssessmentForm(data=data, instance=assessment, prefix="assessment_form")
        full_text_assessment_form = FullTextAssessmentForm(data=data, instance=assessment, prefix="full_text_assessment_form")
        # Get relevance for context.
        if assessment.is_relevant == True:
            is_relevant = 'True'
        else:
            is_relevant = 'False'
        if assessment.full_text_is_relevant == True:
            full_text_is_relevant = 'True'
        elif assessment.full_text_is_relevant == False:
            full_text_is_relevant = 'False'
        else:
            full_text_is_relevant = ''  # Not yet assessed
    else:
        assessment_form = AssessmentForm(data=data, prefix="assessment_form")
        full_text_assessment_form = FullTextAssessmentForm(data=data, prefix="full_text_assessment_form")
        is_relevant = ''  # Not yet assessed
        full_text_is_relevant = ''  # Not yet assessed

    # Formset for this publication
    ExperimentFormSet = modelformset_factory(Experiment, form=ExperimentForm1, extra=4, can_delete=True)
    formset = ExperimentFormSet(data=data, queryset=Experiment.objects.filter(publication=publication, user=user), prefix="experiment_formset")

    if request.method == 'POST':
        if 'is_relevant' in request.POST:
            with transaction.atomic():
                # Update assessment
                if assessment_form.is_valid():
                    assessment = assessment_form.save(commit=False)
                    assessment.user = user
                    assessment.publication = publication
                    assessment.subject = subject
                    assessment.is_relevant = True
                    assessment.is_completed = False
                    assessment.full_text_is_relevant = None
                    assessment.cannot_find = False
                    assessment.cannot_access = False
                    assessment.language_barrier = False
                    assessment.secondary_literature = False
                    assessment.no_population = False
                    assessment.no_intervention = False
                    assessment.no_outcome = False
                    assessment.no_comparator = False
                    assessment.other = False
                    assessment.note = ''
                    assessment.save()
                    # Update status and get next assessment
                    if publication_pk not in completed_assessments:
                        completed_assessments.append(publication_pk)
                        item.completed_assessments = completed_assessments
                    next_assessment = get_next_assessment(publication_pk, next_pk, assessment_order, completed_assessments)
                    item.next_assessment = next_assessment
                    item.save()
                    return redirect('publication', subject=subject.slug, publication_pk=next_assessment)
        if 'is_not_relevant' in request.POST:
            with transaction.atomic():
                if assessment_form.is_valid():
                    # Update assessment
                    assessment = assessment_form.save(commit=False)
                    assessment.user = user
                    assessment.publication = publication
                    assessment.subject = subject
                    assessment.is_relevant = False
                    assessment.is_completed = False
                    assessment.full_text_is_relevant = None
                    assessment.cannot_find = False
                    assessment.cannot_access = False
                    assessment.language_barrier = False
                    assessment.secondary_literature = False
                    assessment.no_population = False
                    assessment.no_intervention = False
                    assessment.no_outcome = False
                    assessment.no_comparator = False
                    assessment.other = False
                    assessment.note = ''
                    assessment.save()
                    # Update status and get next assessment
                    if publication_pk not in completed_assessments:
                        completed_assessments.append(publication_pk)
                        item.completed_assessments = completed_assessments
                    next_assessment = get_next_assessment(publication_pk, next_pk, assessment_order, completed_assessments)
                    item.next_assessment = next_assessment
                    item.save()
                    return redirect('publication', subject=subject.slug, publication_pk=next_assessment)
        if 'full_text_is_not_relevant' in request.POST:
            with transaction.atomic():
                if full_text_assessment_form.is_valid():
                    # Update assessments (add the reason for rejection)
                    assessment = full_text_assessment_form.save(commit=False)
                    assessment.user = user
                    assessment.publication = publication
                    assessment.subject = subject
                    assessment.is_relevant = True  # It must be relevant based on the title and abstract if it is to be rejected based on the full text.
                    assessment.is_completed = True
                    assessment.full_text_is_relevant = False
                    assessment.save()
                    # Update status and get next assessment
                    if publication_pk not in completed_assessments:
                        completed_assessments.append(publication_pk)
                        item.completed_assessments = completed_assessments
                    next_assessment = get_next_assessment(publication_pk, next_pk, assessment_order, completed_assessments)
                    item.next_assessment = next_assessment
                    item.save()
                return redirect('publication', subject=subject.slug, publication_pk=publication_pk)
        if 'reset' in request.POST:
            if publication_pk in completed_assessments:
                completed_assessments.remove(publication_pk)
                item.completed_assessments = completed_assessments
            next_assessment = publication_pk
            item.next_assessment = next_assessment
            item.save()
            if Assessment.objects.filter(publication=publication, user=user).exists():
                Assessment.objects.filter(publication=publication, user=user).delete()
            return redirect('publication', subject=subject.slug, publication_pk=publication_pk)
        if 'save' in request.POST or 'delete' in request.POST:
            with transaction.atomic():
                # Before the formset is validated, the choices need to be redefined, or the validation will fail. This is because only a subset of all choices (high level choices in the MPTT tree) were initially shown in the dropdown (for better UI).
                interventions = TreeNodeChoiceField(queryset=Intervention.objects.filter(pk=intervention.pk).get_descendants(include_self=True), level_indicator = "---")
                for form in formset:
                    form.fields['intervention'] = interventions
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.publication = publication
                            instance.user = user
                            instance.save()
                    if full_text_assessment_form.is_valid():
                        # If an intervention has been selected, mark this publication as "relevant" to this systematic map.
                        # Check if this publication was already marked as "completed" before saving the form. If it was, mark it as not "completed" (because a new intervention has been added, and the user will need to add new metadata).
                        if Experiment.objects.filter(publication=publication, user=user).exists():
                            if Assessment.objects.filter(publication=publication, user=user).exists():
                                old_assessment = Assessment.objects.get(publication=publication, user=user)
                                was_completed = old_assessment.is_completed
                            else:
                                was_completed = False
                            # Get the form data for the new assessment.
                            assessment = full_text_assessment_form.save(commit=False)
                            if was_completed == True:
                                assessment.is_completed = False
                            else:
                                assessment.is_completed = assessment.is_completed
                            assessment.user = user
                            assessment.publication = publication
                            assessment.subject = subject
                            assessment.is_relevant = True
                            assessment.full_text_is_relevant = True
                            assessment.cannot_find = False
                            assessment.cannot_access = False
                            assessment.language_barrier = False
                            assessment.secondary_literature = False
                            assessment.no_population = False
                            assessment.no_intervention = False
                            assessment.no_outcome = False
                            assessment.no_comparator = False
                            assessment.other = False
                            assessment.note = ''
                            assessment.save()
                            # Update status and get next assessment
                            if publication_pk not in completed_assessments:
                                completed_assessments.append(publication_pk)
                                item.completed_assessments = completed_assessments
                            next_assessment = get_next_assessment(publication_pk, next_pk, assessment_order, completed_assessments)
                            item.next_assessment = next_assessment
                            item.save()
                        # If all interventions have been deleted, remove this publication from the completed assessments.
                        elif 'delete' in request.POST:
                            if publication_pk in completed_assessments:
                                completed_assessments.remove(publication_pk)
                                item.completed_assessments = completed_assessments
                            next_assessment = publication_pk
                            item.next_assessment = next_assessment
                            item.save()
                            if Assessment.objects.filter(publication=publication, user=user).exists():
                                Assessment.objects.filter(publication=publication, user=user).delete()
                return redirect('publication', subject=subject.slug, publication_pk=publication_pk)
        if 'set_assessment_order' in request.POST:
            with transaction.atomic():
                form = user_subject_form_1
                if form.is_valid():
                    instance = form.save(commit=False)
                    instance.user = user
                    instance.subject = subject
                    instance.save()
                    user_for_comparison = instance.user_for_comparison
                    if user != user_for_comparison:
                        assessment_status_of_user = AssessmentStatus.objects.get(user=user, subject=subject)
                        assessment_status_of_user_for_comparison = AssessmentStatus.objects.get(user=user_for_comparison, subject=subject)
                        completed_assessments_of_user = literal_eval(assessment_status_of_user.completed_assessments)
                        completed_assessments_of_user_for_comparison = literal_eval(assessment_status_of_user_for_comparison.completed_assessments)
                        if completed_assessments_of_user_for_comparison:  # If the user_for_comparison has completed any assessments
                            assessments_for_comparison = list(set(completed_assessments_of_user_for_comparison) - set(completed_assessments_of_user))
                            shuffle(assessments_for_comparison)
                            new_assessment_order = completed_assessments_of_user + assessments_for_comparison
                            publication_pks = Publication.objects.filter(subject=subject).values_list('pk', flat=True)
                            publication_pks = list(publication_pks)
                            new_assessment_order = new_assessment_order + list(set(publication_pks) - set(new_assessment_order))
                            assessment_status_of_user.assessment_order = new_assessment_order
                            next_assessment = assessments_for_comparison[0]
                            assessment_status_of_user.next_assessment = next_assessment
                            assessment_status_of_user.save()
                            return redirect('publication', subject=subject.slug, publication_pk=next_assessment)
        if 'next_for_kappa' in request.POST or 'previous_for_kappa' in request.POST:
            form = user_subject_form_2
            if form.is_valid():
                instance = form.save(commit=False)
                instance.user = user
                instance.subject = subject
                instance.save()
            if 'next_for_kappa' in request.POST:
                return redirect('full_text_navigation', subject=subject.slug, direction='next', state='kappa', users='this_user', publication_pk=publication_pk)
            if 'previous_for_kappa' in request.POST:
                return redirect('full_text_navigation', subject=subject.slug, direction='previous', state='kappa', users='this_user', publication_pk=publication_pk)
    else:
        # Intervention choices for the formset (high-level choices only)
        interventions = TreeNodeChoiceField(required=False, queryset=Intervention.objects.filter(pk=intervention.pk).get_descendants(include_self=True).filter(level__lte=2).filter(level__gt=0), level_indicator = "---")
        for form in formset:
            form.fields['intervention'] = interventions
    context = {
        'subject': subject,
        'publication': publication,
        'assessment_form': assessment_form,
        'full_text_assessment_form': full_text_assessment_form,
        'user_subject_form_1': user_subject_form_1,
        'user_subject_form_2': user_subject_form_2,
        'experiment_formset': formset,
        'is_relevant': is_relevant,
        'full_text_is_relevant': full_text_is_relevant,
        'next_pk': next_pk,
        'previous_pk': previous_pk
    }
    status = get_status(user, subject)
    context.update(status)
    return render(request, 'publications/publication.html', context)


@login_required
def metadata(request, subject, publication_pk):
    """
    On this page, the user edits the publication-level metadata for this publication (i.e. metadata that is applicable to all interventions).
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    outcome = subject.outcome  # The root outcome for this subject (each subject can have its own classification of outcomes)
    attribute = subject.attribute  # The root attribute for this subject (each subject can have its own classification of attributes)
    attributes = Attribute.objects.get(pk=attribute.pk).get_children().order_by('attribute')
    attributes_count = attributes.count()
    # This publication
    publication = get_object_or_404(Publication, pk=publication_pk, subject=subject)
    # Formsets
    DateFormSet = modelformset_factory(Date, form=DateForm, extra=1, max_num=1, can_delete=True)
    PublicationPopulationFormSet = modelformset_factory(PublicationPopulation, form=PublicationPopulationForm, extra=4, can_delete=True)
    CoordinatesFormSet = modelformset_factory(Coordinates, form=CoordinatesForm, extra=1, can_delete=True)
    EAVFormSet = modelformset_factory(EAV, form=EAVPublicationForm, extra=attributes_count, can_delete=True)
    XCountryFormSet = modelformset_factory(XCountry, form=XCountryForm, extra=1, can_delete=True)
    # Formsets for this publication
    date_formset = DateFormSet(data=data, queryset=Date.objects.filter(publication=publication, user=user), prefix="date_formset")
    coordinates_formset = CoordinatesFormSet(data=data, queryset=Coordinates.objects.filter(publication=publication, user=user), prefix="coordinates_formset")
    x_country_formset = XCountryFormSet(data=data, queryset=XCountry.objects.filter(publication=publication, user=user), prefix="x_country_formset")
    # publication_population_formset
    publication_population_formset = PublicationPopulationFormSet(data=data, queryset=PublicationPopulation.objects.filter(publication=publication, user=user), prefix="publication_population_formset")
    populations = TreeNodeChoiceField(queryset=Outcome.objects.filter(pk=outcome.pk).get_descendants(include_self=True).filter(level=1), level_indicator = "---")
    for form in publication_population_formset:
        form.fields['population'] = populations
    # EAV_formset
    EAV_formset = EAVFormSet(data=data, queryset=EAV.objects.filter(publication=publication, user=user), prefix="EAV_formset")
    for form in EAV_formset:
        # Each form can have a different attribute.
        # If the form has an instance, get the attribute for that instance.
        if form.instance.pk is not None:
            attribute = form.instance.attribute
            form.unit = attribute.unit
        # There should be one "extra" form for each attribute in attributes.
        # If the form does not have an instance, get an attribute, and then
        # delete that attribute from attributes.
        else:
            attribute = attributes[0]
            attributes = attributes.exclude(attribute=attribute)
            form.initial['attribute'] = attribute
            form.initial['publication'] = publication  # For unique_together validation
            form.initial['user'] = user  # For unique_together validation
            form.unit = attribute.unit
        if attribute.is_leaf_node():  # If factor options have not been defined, or if the data type is a number, not a factor
            form.fields['value_as_factor'].disabled = True
        else:  # If factor options have been defined (which is not possible if the data type is a number)
            form.fields['value_as_number'].disabled = True
            form.fields['value_as_factor'] = TreeNodeChoiceField(queryset=attribute.get_children(), level_indicator="")
    if request.method == 'POST':
        if 'save' in request.POST or 'delete' in request.POST:
            with transaction.atomic():
                formset = EAV_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.publication = publication
                            instance.user = user
                            instance.publication_index = publication
                            instance.save()
                formset = x_country_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.publication = publication
                            instance.publication_index = publication
                            instance.user = user
                            instance.save()
                formset = date_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.publication = publication
                            instance.publication_index = publication
                            instance.user = user
                            instance.save()
                formset = coordinates_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.publication = publication
                            instance.publication_index = publication
                            instance.user = user
                            instance.save()
                formset = publication_population_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.publication = publication
                            instance.user = user
                            instance.save()
                return redirect('metadata', subject=subject.slug, publication_pk=publication_pk)
    context = {
        'user_can_edit_attributes': user_subject.can_edit_attributes,
        'subject': subject,
        'publication': publication,
        'publication_population_formset': publication_population_formset,
        'x_country_formset': x_country_formset,
        'coordinates_formset': coordinates_formset,
        'date_formset': date_formset,
        'EAV_formset': EAV_formset
    }
    return render(request, 'publications/metadata.html', context)


@login_required
def attributes(request, subject):
    """
    On this page, the user adds/edits attributes for this subject.
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject, can_edit_attributes=True)  # Check if this user has permission to work on this subject.
    attribute = subject.attribute  # The root attribute for this subject (each subject can have its own classification of attributes)
    attributes = Attribute.objects.get(pk=attribute.pk).get_children()
    FormSet = modelformset_factory(Attribute, form=AttributeForm, extra=1, can_delete=True)
    formset = FormSet(data=data, queryset=attributes, initial=[{'parent': attribute}])  # To check for unique_together = ('attribute', 'parent'), parent needs to be passed to the formset as initial data, but it is hidden in the template and ignored in the view.
    if request.method == 'POST':
        if 'save' in request.POST or 'delete' in request.POST:
            with transaction.atomic():
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.parent = attribute
                            instance.user = user
                            if instance.type == 'factor':
                                instance.unit = 'NA for factors'
                            instance.save()
                    return redirect('attributes', subject=subject.slug)
    context = {
        'subject': subject,
        'formset': formset
    }
    return render(request, 'publications/attributes.html', context)


@login_required
def attribute(request, subject, attribute_pk):
    """
    On this page, the user adds/edits options for this attribute.
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject, can_edit_attributes=True)  # Check if this user has permission to work on this subject.
    attribute = Attribute.objects.get(pk=attribute_pk)
    options = attribute.get_children()
    FormSet = modelformset_factory(Attribute, form=AttributeOptionForm, extra=1, can_delete=True)
    formset = FormSet(data=data, queryset=options, initial=[{'parent': attribute}])  # To check for unique_together = ('attribute', 'parent'), parent needs to be passed to the formset as initial data, but it is hidden in the template and ignored in the view.
    if request.method == 'POST':
        if 'save' in request.POST or 'delete' in request.POST:
            with transaction.atomic():
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.parent = attribute
                            instance.user = user
                            instance.type = 'factor'
                            instance.save()
                return redirect('attribute', subject=subject.slug, attribute_pk=attribute_pk)
    context = {
        'subject': subject,
        'formset': formset,
        'attribute': attribute
    }
    return render(request, 'publications/attribute.html', context)


@login_required
def publication_population(request, subject, publication_pk, publication_population_index):
    """
    On this page, the user chooses an outcome for this publication population (not for a specific intervention).
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    outcome = subject.outcome  # The root outcome for this subject (each subject can have its own classification of outcomes)
    PublicationPopulationOutcomeFormSet = modelformset_factory(PublicationPopulationOutcome, form=PublicationPopulationOutcomeForm, extra=4, can_delete=True)
    # This publication
    publication = get_object_or_404(Publication, pk=publication_pk, subject=subject)
    # This publication_population
    publication_populations = PublicationPopulation.objects.filter(publication=publication, user=user).order_by('pk')
    publication_population = publication_populations[publication_population_index]
    # Formset for this publication_population
    formset = PublicationPopulationOutcomeFormSet(data=data, queryset=PublicationPopulationOutcome.objects.filter(publication_population=publication_population), prefix="publication_population_outcome_formset")
    # Outcome choices for the formset
    for form in formset:
        form.fields['outcome'] = TreeNodeChoiceField(queryset=Outcome.objects.get(pk=publication_population.population.pk).get_descendants(include_self=True), level_indicator = "---")
    if request.method == 'POST':
        with transaction.atomic():
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.publication_population = publication_population
                        instance.user = user
                        instance.save()
            return redirect('publication_population', subject=subject.slug, publication_pk=publication_pk, publication_population_index=publication_population_index)
    context = {
        'subject': subject,
        'publication': publication,
        'publication_population': publication_population,
        'publication_population_index': publication_population_index,
        'formset': formset
    }
    return render(request, 'publications/publication_population.html', context)


@login_required
def add_publication(request, subject):
    """
    On this page, the user adds a publication that has not come from a systematic search (e.g., publications from "other sources" in a PRISMA diagram).
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    publication_form = PublicationForm(data=data)
    if request.method == 'POST':
        with transaction.atomic():
            form = publication_form
            if form.is_valid():
                with reversion.create_revision():  # Version control (django-revision)
                    instance = form.save(commit=False)
                    instance.subject = subject
                    instance.is_from_systematic_search = False
                    instance.save()
                    publication_pk = instance.pk
                    reversion.set_user(request.user)
        return redirect('publication', subject=subject.slug, publication_pk=publication_pk)
    context = {
        'subject': subject,
        'publication_form': publication_form
    }
    # Get data for the sidebar.
    status = get_status(user, subject)
    context.update(status)
    return render(request, 'publications/add_publication.html', context)


@login_required
def edit_publication(request, subject, publication_pk):
    """
    On this page, the user edits the title, abstract, etc. for this publication.
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    publication_pk = int(publication_pk)
    # This publication
    publication = get_object_or_404(Publication, pk=publication_pk, subject=subject)
    publication_form = PublicationForm(data=data, instance=publication)
    if request.method == 'POST':
        publication_form = PublicationForm(request.POST, instance=publication)
        if publication_form.is_valid():
            with reversion.create_revision():  # Version control (django-revision)
                publication_form.save()
                reversion.set_user(request.user)
        return redirect('publication', subject=subject.slug, publication_pk=publication_pk)
    context = {
        'subject': subject,
        'publication': publication,
        'publication_form': publication_form
    }
    # Get data for the sidebar.
    status = get_status(user, subject)
    context.update(status)
    return render(request, 'publications/edit_publication.html', context)


@login_required
def experiment(request, subject, publication_pk, experiment_index):
    """
    On this page, the user chooses populations, experimental designs, and tags for this intervention (i.e. this "experiment").
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    design = subject.design  # The root design for this subject (each subject can have its own classification of designs)
    intervention = subject.intervention  # The root intervention for this subject (each subject can have its own classification of interventions)
    outcome = subject.outcome  # The root outcome for this subject (each subject can have its own classification of outcomes)
    attribute = subject.attribute  # The root attribute for this subject (each subject can have its own classification of attributes)
    attributes = Attribute.objects.get(pk=attribute.pk).get_children().order_by('attribute')
    attributes_count = attributes.count()
    # This publication
    publication = get_object_or_404(Publication, pk=publication_pk, subject=subject)
    # This experiment
    experiments = Experiment.objects.filter(publication=publication, user=user).order_by('pk')
    experiment = experiments[experiment_index]
    # Forms
    experiment_form = ExperimentForm2(data=data, instance=experiment, prefix="experiment_form")
    experiment_form.fields['intervention'] = TreeNodeChoiceField(queryset=Intervention.objects.filter(pk=intervention.pk).get_descendants(include_self=True), level_indicator = "---")
    # Formsets
    ExperimentDesignFormSet = modelformset_factory(ExperimentDesign, form=ExperimentDesignForm, extra=5, max_num=5, can_delete=True)
    experiment_design_formset = ExperimentDesignFormSet(data=data, queryset=ExperimentDesign.objects.filter(experiment=experiment), prefix="experiment_design_formset")
    designs = TreeNodeChoiceField(required=False, queryset=Design.objects.filter(pk=design.pk).get_descendants(include_self=True).filter(level__gte=1), level_indicator = "---")
    for form in experiment_design_formset:
        form.fields['design'] = designs
    ExperimentPopulationFormSet = modelformset_factory(ExperimentPopulation, form=ExperimentPopulationForm, extra=4, can_delete=True)
    experiment_population_formset = ExperimentPopulationFormSet(data=data, queryset=ExperimentPopulation.objects.filter(experiment=experiment), prefix="experiment_population_formset")
    populations = TreeNodeChoiceField(required=False, queryset=Outcome.objects.filter(pk=outcome.pk).get_descendants(include_self=True).filter(level=1), level_indicator = "---")
    for form in experiment_population_formset:
        form.fields['population'] = populations
    CoordinatesFormSet = modelformset_factory(Coordinates, form=CoordinatesForm, extra=1, can_delete=True)
    coordinates_formset = CoordinatesFormSet(data=data, queryset=Coordinates.objects.filter(experiment=experiment), prefix="coordinates_formset")
    DateFormSet = modelformset_factory(Date, form=DateForm, extra=1, max_num=1, can_delete=True)
    date_formset = DateFormSet(data=data, queryset=Date.objects.filter(experiment=experiment), prefix="date_formset")
    StudyFormSet = modelformset_factory(Study, form=StudyForm, extra=1, max_num=1, can_delete=True)
    study_formset = StudyFormSet(data=data, queryset=Study.objects.filter(experiment=experiment), prefix="study_formset")
    XCountryFormSet = modelformset_factory(XCountry, form=XCountryForm, extra=1, can_delete=True)
    x_country_formset = XCountryFormSet(data=data, queryset=XCountry.objects.filter(experiment=experiment), prefix="x_country_formset")
    EAVFormSet = modelformset_factory(EAV, form=EAVExperimentForm, extra=attributes_count, can_delete=True)
    EAV_formset = EAVFormSet(data=data, queryset=EAV.objects.filter(experiment=experiment), prefix="EAV_formset")
    for form in EAV_formset:
        # Each form can have a different attribute.
        # If the form has an instance, get the attribute for that instance.
        if form.instance.pk is not None:
            attribute = form.instance.attribute
            form.unit = attribute.unit
        # There should be one "extra" form for each attribute in attributes.
        # If the form does not have an instance, get an attribute, and then
        # delete that attribute from attributes.
        else:
            attribute = attributes[0]
            attributes = attributes.exclude(attribute=attribute)
            form.initial['attribute'] = attribute
            form.initial['experiment'] = experiment  # For unique_together validation
            form.initial['user'] = user  # For unique_together validation
            form.unit = attribute.unit
        if attribute.is_leaf_node():  # If factor options have not been defined, or if the data type is a number, not a factor
            form.fields['value_as_factor'].disabled = True
        else:  # If factor options have been defined (which is not possible if the data type is a number)
            form.fields['value_as_number'].disabled = True
            form.fields['value_as_factor'] = TreeNodeChoiceField(queryset=attribute.get_children(), level_indicator="")
    if request.method == 'POST':
        with transaction.atomic():
            form = experiment_form
            if form.is_valid():
                form.save()
            formset = experiment_population_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.experiment = experiment
                        instance.save()
            formset = experiment_design_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.experiment = experiment
                        instance.save()
            formset = x_country_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.experiment = experiment
                        instance.experiment_index = experiment
                        instance.publication_index = publication
                        instance.user = user
                        instance.save()
            formset = coordinates_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.experiment = experiment
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.user = user
                        instance.save()
            formset = date_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.experiment = experiment
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.user = user
                        instance.save()
            formset = study_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.experiment = experiment
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.user = user
                        instance.save()
            formset = EAV_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.experiment = experiment
                        instance.user = user
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.save()
            return redirect('experiment', subject=subject.slug, publication_pk=publication_pk, experiment_index=experiment_index)
    context = {
        'user_can_edit_attributes': user_subject.can_edit_attributes,
        'subject': subject,
        'publication': publication,
        'experiment': experiment,
        'experiment_index': experiment_index,
        'experiment_form': experiment_form,
        'experiment_design_formset': experiment_design_formset,
        'experiment_population_formset': experiment_population_formset,
        'coordinates_formset': coordinates_formset,
        'date_formset': date_formset,
        'study_formset': study_formset,
        'EAV_formset': EAV_formset,
        'x_country_formset': x_country_formset
    }
    return render(request, 'publications/experiment.html', context)


@login_required
def population(request, subject, publication_pk, experiment_index, population_index):
    """
    On this page, the user chooses outcomes for this population.
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    attribute = subject.attribute  # The root attribute for this subject (each subject can have its own classification of attributes)
    attributes = Attribute.objects.get(pk=attribute.pk).get_children().order_by('attribute')
    attributes_count = attributes.count()
    # This publication
    publication = get_object_or_404(Publication, pk=publication_pk, subject=subject)
    # This experiment
    experiments = Experiment.objects.filter(publication=publication, user=user).order_by('pk')
    experiment = experiments[experiment_index]
    # This population
    experiment_populations = ExperimentPopulation.objects.filter(experiment=experiment).order_by('pk')
    experiment_population = experiment_populations[population_index]
    data_upload_form = DataUploadForm(request.POST, request.FILES)
    # Formsets
    ExperimentPopulationOutcomeFormSet = modelformset_factory(ExperimentPopulationOutcome, form=ExperimentPopulationOutcomeForm, extra=4, can_delete=True)
    experiment_population_outcome_formset = ExperimentPopulationOutcomeFormSet(data=data, queryset=ExperimentPopulationOutcome.objects.filter(experiment_population=experiment_population), prefix="experiment_population_outcome_formset")
    outcomes = TreeNodeChoiceField(required=False, queryset=Outcome.objects.get(pk=experiment_population.population.pk).get_descendants(include_self=True), level_indicator = "---")
    for form in experiment_population_outcome_formset:
        form.fields['outcome'] = outcomes
    CoordinatesFormSet = modelformset_factory(Coordinates, form=CoordinatesForm, extra=1, can_delete=True)
    coordinates_formset = CoordinatesFormSet(data=data, queryset=Coordinates.objects.filter(population=experiment_population), prefix="coordinates_formset")
    DateFormSet = modelformset_factory(Date, form=DateForm, extra=1, max_num=1, can_delete=True)
    date_formset = DateFormSet(data=data, queryset=Date.objects.filter(population=experiment_population), prefix="date_formset")
    StudyFormSet = modelformset_factory(Study, form=StudyForm, extra=1, max_num=1, can_delete=True)
    study_formset = StudyFormSet(data=data, queryset=Study.objects.filter(population=experiment_population), prefix="study_formset")
    XCountryFormSet = modelformset_factory(XCountry, form=XCountryForm, extra=1, can_delete=True)
    x_country_formset = XCountryFormSet(data=data, queryset=XCountry.objects.filter(population=experiment_population), prefix="x_country_formset")
    EAVFormSet = modelformset_factory(EAV, form=EAVExperimentForm, extra=attributes_count, can_delete=True)
    EAV_formset = EAVFormSet(data=data, queryset=EAV.objects.filter(population=experiment_population), prefix="EAV_formset")
    for form in EAV_formset:
        # Each form can have a different attribute.
        # If the form has an instance, get the attribute for that instance.
        if form.instance.pk is not None:
            attribute = form.instance.attribute
            form.unit = attribute.unit
        # There should be one "extra" form for each attribute in attributes.
        # If the form does not have an instance, get an attribute, and then
        # delete that attribute from attributes.
        else:
            attribute = attributes[0]
            attributes = attributes.exclude(attribute=attribute)
            form.initial['attribute'] = attribute
            form.initial['population'] = experiment_population  # For unique_together validation
            form.initial['user'] = user  # For unique_together validation
            form.unit = attribute.unit
        if attribute.is_leaf_node():  # If factor options have not been defined, or if the data type is a number, not a factor
            form.fields['value_as_factor'].disabled = True
        else:  # If factor options have been defined (which is not possible if the data type is a number)
            form.fields['value_as_number'].disabled = True
            form.fields['value_as_factor'] = TreeNodeChoiceField(queryset=attribute.get_children(), level_indicator="")
    # Columns for file upload template
    col_names = ['outcome', 'comparison', 'treatment_mean', 'control_mean', 'treatment_sd', 'control_sd', 'treatment_n', 'control_n', 'treatment_se', 'control_se', 'treatment_mean_before', 'control_mean_before', 'treatment_sd_before', 'control_sd_before', 'treatment_n_before', 'control_n_before', 'treatment_se_before', 'control_se_before', 'unit', 'lsd', 'is_significant', 'approximate_p_value', 'p_value', 'z_value', 'n', 'correlation_coefficient', 'note', 'country']
    study_cols = ['study_id', 'study_name']
    coordinates_cols = ['latitude_degrees', 'latitude_minutes', 'latitude_seconds', 'latitude_direction', 'longitude_degrees', 'longitude_minutes', 'longitude_seconds', 'longitude_direction']
    date_cols = ['start_year', 'start_month', 'start_day', 'end_year', 'end_month', 'end_day']
    col_names = col_names + coordinates_cols
    col_names = col_names + date_cols
    col_names = col_names + study_cols
    integer_cols = ['treatment_n', 'control_n', 'treatment_n_before', 'control_n_before', 'n', 'study_id', 'start_year', 'end_year']
    decimal_cols = ['treatment_mean', 'control_mean', 'treatment_sd', 'control_sd', 'treatment_se', 'control_se', 'treatment_mean_before', 'control_mean_before', 'treatment_sd_before', 'control_sd_before', 'treatment_se_before', 'control_se_before', 'lsd', 'z_value']
    minutes_seconds_cols = ['latitude_minutes', 'latitude_seconds', 'longitude_minutes', 'longitude_seconds']
    year_cols = ['start_year', 'end_year']
    month_cols = ['start_month', 'end_month']
    day_cols = ['start_day', 'end_day']
    attribute = subject.attribute
    attributes = Attribute.objects.get(pk=subject.attribute.pk).get_children().order_by('attribute')
    for attribute in attributes.values_list('attribute', flat=True):
        col_names.append(attribute)
    col_names_dict = {}

    if request.method == 'POST':

        if 'save' in request.POST or 'delete' in request.POST:
            with transaction.atomic():
                formset = experiment_population_outcome_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.experiment_population = experiment_population
                            instance.save()
                formset = x_country_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.population = experiment_population
                            instance.user = user
                            instance.publication_index = publication
                            instance.experiment_index = experiment
                            instance.population_index = experiment_population
                            instance.save()
                formset = coordinates_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.population = experiment_population
                            instance.user = user
                            instance.publication_index = publication
                            instance.experiment_index = experiment
                            instance.population_index = experiment_population
                            instance.save()
                formset = date_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.population = experiment_population
                            instance.user = user
                            instance.publication_index = publication
                            instance.experiment_index = experiment
                            instance.population_index = experiment_population
                            instance.save()
                formset = study_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.population = experiment_population
                            instance.user = user
                            instance.publication_index = publication
                            instance.experiment_index = experiment
                            instance.population_index = experiment_population
                            instance.save()
                formset = EAV_formset
                if formset.is_valid():
                    instances = formset.save(commit=False)
                    if 'delete' in request.POST:
                        for obj in formset.deleted_objects:
                            obj.delete()
                    else:
                        for instance in instances:
                            instance.population = experiment_population
                            instance.user = user
                            instance.publication_index = publication
                            instance.experiment_index = experiment
                            instance.population_index = experiment_population
                            instance.save()

        if 'upload' in request.POST:
            attributes = Attribute.objects.get(pk=subject.attribute.pk).get_children()
            outcomes = Outcome.objects.get(pk=subject.outcome.pk).get_descendants(include_self=True)
            form = data_upload_form
            if form.is_valid():
                file = form.cleaned_data.get('data_upload_file')
                wb = load_workbook(file)
                ws = wb["Data"]
                col_names = ws[1]
                with transaction.atomic():
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value:  # If this row has an outcome
                            for cell in row:
                                if (cell.value is not None or cell.value == 0):
                                    attribute = col_names[cell.column - 1].value
                                    if attribute == "outcome":
                                        outcome = cell.value
                                        outcome = outcomes.filter(outcome=outcome).order_by('-level')[0]
                                        experiment_population_outcome = ExperimentPopulationOutcome(
                                            experiment_population = experiment_population,
                                            outcome = outcome
                                        )
                                        experiment_population_outcome.save()
                                        # For models with data from multiple cells, create instances to be saved later.
                                        data = Data(
                                            subject = subject,
                                            publication = publication,
                                            experiment = experiment,
                                            experiment_population = experiment_population,
                                            experiment_population_outcome = experiment_population_outcome
                                        )
                                        coordinates = Coordinates(
                                            user = user,
                                            outcome = experiment_population_outcome,
                                            publication_index = publication,
                                            experiment_index = experiment,
                                            population_index = experiment_population,
                                            outcome_index = experiment_population_outcome
                                        )
                                        date = Date(
                                            user = user,
                                            outcome = experiment_population_outcome,
                                            publication_index = publication,
                                            experiment_index = experiment,
                                            population_index = experiment_population,
                                            outcome_index = experiment_population_outcome
                                        )
                                        study = Study(
                                            user = user,
                                            outcome = experiment_population_outcome,
                                            publication_index = publication,
                                            experiment_index = experiment,
                                            population_index = experiment_population,
                                            outcome_index = experiment_population_outcome
                                        )
                                    elif attribute == "country":
                                        country = cell.value
                                        country = Country.objects.get(country=country)
                                        x_country = XCountry(
                                            country = country,
                                            user = user,
                                            outcome = experiment_population_outcome,
                                            publication_index = publication,
                                            experiment_index = experiment,
                                            population_index = experiment_population,
                                            outcome_index = experiment_population_outcome
                                        )
                                        x_country.save()
                                    elif attribute in attributes.values_list('attribute', flat=True):
                                        attribute = Attribute.objects.get(attribute=attribute)
                                        eav = EAV(
                                            attribute = attribute,
                                            user = user,
                                            outcome = experiment_population_outcome,
                                            publication_index = publication,
                                            experiment_index = experiment,
                                            population_index = experiment_population,
                                            outcome_index = experiment_population_outcome
                                        )
                                        if attribute.type == "factor":
                                            value = attribute.get_children().get(attribute=cell.value)
                                            eav.value_as_factor = value
                                        elif attribute.type == "number":
                                            eav.value_as_number = cell.value
                                        eav.save()
                                    # For models with data from multiple cells
                                    elif attribute in coordinates_cols:
                                        if attribute == "latitude_direction" or attribute == "longitude_direction":
                                            cell.value = cell.value.upper()
                                        setattr(coordinates, attribute, cell.value)
                                    elif attribute in date_cols:
                                        setattr(date, attribute, cell.value)
                                    elif attribute in study_cols:
                                        setattr(study, attribute, cell.value)
                                    else:
                                        setattr(data, attribute, cell.value)
                            # For models with data from multiple cells, save the model instance only if data has been entered.
                            # If data has been entered for the Data instance, save the instance (here we require both treatment_mean and control_mean, but this constraint is not enforced by the Data model, in case there are publications that report an effect size but not a treatment mean and control mean).
                            if data.treatment_mean == None or data.control_mean == None:
                                pass
                            else:
                                data.save()
                            # If data has been entered for the Study instance, save it.
                            if study.study_name == None and study.study_id == None:
                                pass
                            else:
                                study.save()
                            # If data has been entered for the Coordinates instance, save it.
                            if coordinates.latitude_degrees == None and coordinates.latitude_minutes == None and coordinates.latitude_seconds == None and coordinates.longitude_degrees == None and coordinates.longitude_minutes == None and coordinates.longitude_seconds == None:
                                pass
                            else:
                                coordinates.save()
                            # If data has been entered for the Date instance, save it.
                            if date.start_year == None and date.start_month == None and date.start_day == None:
                                pass
                            else:
                                date.save()
                        else:  # If this row does not have an outcome, create a new data point for the previous outcome (but do not create new covariates, since all data points for one outcome must have the same covariates).
                            data = Data(
                                subject = subject,
                                publication = publication,
                                experiment = experiment,
                                experiment_population = experiment_population,
                                experiment_population_outcome = experiment_population_outcome
                            )
                            for cell in row:
                                if (cell.value is not None or cell.value == 0):
                                    attribute = col_names[cell.column - 1].value
                                    if attribute not in attributes.values_list('attribute', flat=True):
                                        setattr(data, attribute, cell.value)
                            # If data has been entered for the Data instance, save the instance (here we require both treatment_mean and control_mean, but this constraint is not enforced by the Data model, in case there are publications that report an effect size but not a treatment mean and control mean).
                            if data.treatment_mean == None or data.control_mean == None:
                                pass
                            else:
                                data.save()
                        # If no data were entered for this outcome, delete this outcome.
                        if not Data.objects.filter(
                            subject = subject,
                            publication = publication,
                            experiment = experiment,
                            experiment_population = experiment_population,
                            experiment_population_outcome = experiment_population_outcome
                        ).exists():
                            experiment_population_outcome.delete()

        elif 'download' in request.POST:
            wb = Workbook()
            ws_1 = wb.active
            ws_1.title = "Data"
            ws_2 = wb.create_sheet("Options")
            i = 1
            for name in col_names:
                col_names_dict[name] = i
                ws_1.cell(row=1, column=i).value = str(name)
                i += 1

            # XLSX data validation for integers
            dv = DataValidation(type="whole", operator="greaterThan", formula1=0)
            dv.error ='Please enter an integer.'
            ws_1.add_data_validation(dv)
            for col in integer_cols:
                column_number = col_names_dict[col]
                column_letter = get_column_letter(column_number)
                dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for decimals
            dv = DataValidation(type="decimal")
            dv.error ='Please enter a number.'
            ws_1.add_data_validation(dv)
            for col in decimal_cols:
                column_number = col_names_dict[col]
                column_letter = get_column_letter(column_number)
                dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for p_value (0 <= p_value <= 1)
            dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
            dv.error ='Please enter a number between 0 and 1.'
            ws_1.add_data_validation(dv)
            column_number = col_names_dict["p_value"]
            column_letter = get_column_letter(column_number)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for correlation_coefficient (-1 <= correlation_coefficient <= 1)
            dv = DataValidation(type="decimal", operator="between", formula1=-1, formula2=1)
            dv.error ='Please enter a number between -1 and 1.'
            ws_1.add_data_validation(dv)
            column_number = col_names_dict["correlation_coefficient"]
            column_letter = get_column_letter(column_number)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for minutes and seconds (0-60)
            dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=60)
            dv.error ='Please enter a number between 0 and 60.'
            ws_1.add_data_validation(dv)
            for col in minutes_seconds_cols:
                column_number = col_names_dict[col]
                column_letter = get_column_letter(column_number)
                dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for latitude_degrees (0-90)
            dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=90)
            dv.error ='Please enter a number between 0 and 90.'
            ws_1.add_data_validation(dv)
            column_number = col_names_dict["latitude_degrees"]
            column_letter = get_column_letter(column_number)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for longitude_degrees (0-180)
            dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=180)
            dv.error ='Please enter a number between 0 and 180.'
            ws_1.add_data_validation(dv)
            column_number = col_names_dict["longitude_degrees"]
            column_letter = get_column_letter(column_number)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for month (1-12)
            dv = DataValidation(type="whole", operator="between", formula1=1, formula2=12)
            dv.error ='Please enter an integer between 1 and 12.'
            ws_1.add_data_validation(dv)
            for col in month_cols:
                column_number = col_names_dict[col]
                column_letter = get_column_letter(column_number)
                dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX data validation for day (1-31)
            dv = DataValidation(type="whole", operator="between", formula1=1, formula2=31)
            dv.error ='Please enter an integer between 1 and 31.'
            ws_1.add_data_validation(dv)
            for col in day_cols:
                column_number = col_names_dict[col]
                column_letter = get_column_letter(column_number)
                dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX options for outcome
            outcomes = Outcome.objects.filter(pk=experiment_population.population.pk).get_descendants(include_self=True)
            column_number = col_names_dict["outcome"]
            column_letter = get_column_letter(column_number)
            options = outcomes.values_list('outcome', flat=True)
            i = 1
            for option in options:
                ws_2.cell(row=i, column=column_number).value = option
                i += 1
            dv = DataValidation(type="list", formula1="=Options!{column_letter}$1:{column_letter}${max_row}".format(column_letter=column_letter, max_row=len(options)))
            dv.error ='Please select an option from the menu.'
            ws_1.add_data_validation(dv)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX options for country
            countries = Country.objects.all()
            column_number = col_names_dict["country"]
            column_letter = get_column_letter(column_number)
            options = countries.values_list('country', flat=True)
            i = 1
            for option in options:
                ws_2.cell(row=i, column=column_number).value = option
                i += 1
            dv = DataValidation(type="list", formula1="=Options!{column_letter}$1:{column_letter}${max_row}".format(column_letter=column_letter, max_row=len(options)))
            dv.error ='Please select an option from the menu.'
            ws_1.add_data_validation(dv)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX options for approximate_p_value
            column_number = col_names_dict["approximate_p_value"]
            column_letter = get_column_letter(column_number)
            options = [
                "< 0.0001",
                "< 0.001",
                "< 0.01",
                "< 0.05",
                "< 0.1",
                "> 0.05",
                "> 0.1",
            ]
            i = 1
            for option in options:
                ws_2.cell(row=i, column=column_number).value = option
                i += 1
            dv = DataValidation(type="list", formula1="=Options!{column_letter}$1:{column_letter}${max_row}".format(column_letter=column_letter, max_row=len(options)))
            ws_1.add_data_validation(dv)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX options for is_significant
            column_number = col_names_dict["is_significant"]
            column_letter = get_column_letter(column_number)
            options = [
                "False",
                "True"
            ]
            i = 1
            for option in options:
                ws_2.cell(row=i, column=column_number).value = option
                i += 1
            dv = DataValidation(type="list", formula1="=Options!{column_letter}$1:{column_letter}${max_row}".format(column_letter=column_letter, max_row=len(options)))
            ws_1.add_data_validation(dv)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX options for latitude_direction
            column_number = col_names_dict["latitude_direction"]
            column_letter = get_column_letter(column_number)
            options = [
                "N",
                "S"
            ]
            i = 1
            for option in options:
                ws_2.cell(row=i, column=column_number).value = option
                i += 1
            dv = DataValidation(type="list", formula1="=Options!{column_letter}$1:{column_letter}${max_row}".format(column_letter=column_letter, max_row=len(options)))
            ws_1.add_data_validation(dv)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX options for longitude_direction
            column_number = col_names_dict["longitude_direction"]
            column_letter = get_column_letter(column_number)
            options = [
                "E",
                "W"
            ]
            i = 1
            for option in options:
                ws_2.cell(row=i, column=column_number).value = option
                i += 1
            dv = DataValidation(type="list", formula1="=Options!{column_letter}$1:{column_letter}${max_row}".format(column_letter=column_letter, max_row=len(options)))
            ws_1.add_data_validation(dv)
            dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # XLSX options for user-defined attributes
            for attribute in attributes:
                column_number = col_names_dict[attribute.attribute]
                column_letter = get_column_letter(column_number)
                if attribute.get_children().exists():
                    options = attribute.get_children().values_list('attribute', flat=True)
                    i = 1
                    for option in options:
                        ws_2.cell(row=i, column=column_number).value = option
                        i += 1
                    dv = DataValidation(type="list", formula1="=Options!{column_letter}$1:{column_letter}${max_row}".format(column_letter=column_letter, max_row=len(options)))
                    dv.error ='Please select an option from the menu.'
                    ws_1.add_data_validation(dv)
                    dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))
                elif attribute.type == "number":
                    dv = DataValidation(type="decimal")
                    dv.error ='Please enter a number.'
                    ws_1.add_data_validation(dv)
                    dv.add("{column_letter}2:{column_letter}201".format(column_letter=column_letter))

            # Create a temporary file for the response
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                response = HttpResponse(content=tmp, content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename=data_upload_template.xlsx'
                return response
            return response
        return redirect('population', subject=subject.slug, publication_pk=publication_pk, experiment_index=experiment_index, population_index=population_index)

    context = {
        'user_can_edit_attributes': user_subject.can_edit_attributes,
        'subject': subject,
        'publication': publication,
        'experiment': experiment,
        'experiment_population': experiment_population,
        'experiment_index': experiment_index,
        'population_index': population_index,
        'experiment_population_outcome_formset': experiment_population_outcome_formset,
        'coordinates_formset': coordinates_formset,
        'date_formset': date_formset,
        'study_formset': study_formset,
        'EAV_formset': EAV_formset,
        'x_country_formset': x_country_formset,
        'data_upload_form': data_upload_form
    }
    return render(request, 'publications/population.html', context)


@login_required
def outcome(request, subject, publication_pk, experiment_index, population_index, outcome_index):
    """
    On this page, the user enters effect sizes and related data for this outcome.
    """
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    attribute = subject.attribute  # The root attribute for this subject (each subject can have its own classification of attributes)
    attributes = Attribute.objects.get(pk=attribute.pk).get_children().order_by('attribute')
    attributes_count = attributes.count()
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    # This publication
    publication = get_object_or_404(Publication, pk=publication_pk, subject=subject)
    # This experiment
    experiments = Experiment.objects.filter(publication=publication, user=user).order_by('pk')
    experiment = experiments[experiment_index]
    # This population
    experiment_populations = ExperimentPopulation.objects.filter(experiment=experiment).order_by('pk')
    experiment_population = experiment_populations[population_index]
    population = experiment_population.population
    # This outcome
    experiment_population_outcomes = ExperimentPopulationOutcome.objects.filter(experiment_population=experiment_population).order_by('pk')
    experiment_population_outcome = experiment_population_outcomes[outcome_index]
    # Forms
    experiment_population_outcome_form = ExperimentPopulationOutcomeForm(data=data, instance=experiment_population_outcome, prefix="experiment_population_outcome_form")
    experiment_population_outcome_form.fields['outcome'] = TreeNodeChoiceField(queryset=Outcome.objects.filter(pk=population.pk).get_descendants(include_self=True), level_indicator = "---")
    # Formsets
    DataFormSet = modelformset_factory(Data, form=DataForm, extra=1, can_delete=True)
    data_formset = DataFormSet(data=data, queryset=Data.objects.filter(experiment_population_outcome=experiment_population_outcome.pk), prefix="data_formset")
    CoordinatesFormSet = modelformset_factory(Coordinates, form=CoordinatesForm, extra=1, can_delete=True)
    coordinates_formset = CoordinatesFormSet(data=data, queryset=Coordinates.objects.filter(outcome=experiment_population_outcome), prefix="coordinates_formset")
    DateFormSet = modelformset_factory(Date, form=DateForm, extra=1, max_num=1, can_delete=True)
    date_formset = DateFormSet(data=data, queryset=Date.objects.filter(outcome=experiment_population_outcome), prefix="date_formset")
    StudyFormSet = modelformset_factory(Study, form=StudyForm, extra=1, max_num=1, can_delete=True)
    study_formset = StudyFormSet(data=data, queryset=Study.objects.filter(outcome=experiment_population_outcome), prefix="study_formset")
    XCountryFormSet = modelformset_factory(XCountry, form=XCountryForm, extra=1, can_delete=True)
    x_country_formset = XCountryFormSet(data=data, queryset=XCountry.objects.filter(outcome=experiment_population_outcome), prefix="x_country_formset")
    EAVFormSet = modelformset_factory(EAV, form=EAVExperimentForm, extra=attributes_count, can_delete=True)
    EAV_formset = EAVFormSet(data=data, queryset=EAV.objects.filter(outcome=experiment_population_outcome), prefix="EAV_formset")
    for form in EAV_formset:
        # Each form can have a different attribute.
        # If the form has an instance, get the attribute for that instance.
        if form.instance.pk is not None:
            attribute = form.instance.attribute
            form.unit = attribute.unit
        # There should be one "extra" form for each attribute in attributes.
        # If the form does not have an instance, get an attribute, and then
        # delete that attribute from attributes.
        else:
            attribute = attributes[0]
            attributes = attributes.exclude(attribute=attribute)
            form.initial['attribute'] = attribute
            form.initial['outcome'] = experiment_population_outcome  # For unique_together validation
            form.initial['user'] = user  # For unique_together validation
            form.unit = attribute.unit
        if attribute.is_leaf_node():  # If factor options have not been defined, or if the data type is a number, not a factor
            form.fields['value_as_factor'].disabled = True
        else:  # If factor options have been defined (which is not possible if the data type is a number)
            form.fields['value_as_number'].disabled = True
            form.fields['value_as_factor'] = TreeNodeChoiceField(queryset=attribute.get_children(), level_indicator="")
    if request.method == 'POST':
        with transaction.atomic():
            form = experiment_population_outcome_form
            if form.is_valid():
                form.save()
            formset = data_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.subject = subject
                        instance.publication = publication
                        instance.experiment = experiment
                        instance.experiment_population = experiment_population
                        instance.experiment_population_outcome = experiment_population_outcome
                        instance.save()
            formset = x_country_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.outcome = experiment_population_outcome
                        instance.user = user
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.population_index = experiment_population
                        instance.outcome_index = experiment_population_outcome
                        instance.save()
            formset = coordinates_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.outcome = experiment_population_outcome
                        instance.user = user
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.population_index = experiment_population
                        instance.outcome_index = experiment_population_outcome
                        instance.save()
            formset = date_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.outcome = experiment_population_outcome
                        instance.user = user
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.population_index = experiment_population
                        instance.outcome_index = experiment_population_outcome
                        instance.save()
            formset = study_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.outcome = experiment_population_outcome
                        instance.user = user
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.population_index = experiment_population
                        instance.outcome_index = experiment_population_outcome
                        instance.save()
            formset = EAV_formset
            if formset.is_valid():
                instances = formset.save(commit=False)
                if 'delete' in request.POST:
                    for obj in formset.deleted_objects:
                        obj.delete()
                else:
                    for instance in instances:
                        instance.outcome = experiment_population_outcome
                        instance.user = user
                        instance.publication_index = publication
                        instance.experiment_index = experiment
                        instance.population_index = experiment_population
                        instance.outcome_index = experiment_population_outcome
                        instance.save()
            return redirect('outcome', subject=subject.slug, publication_pk=publication_pk, experiment_index=experiment_index, population_index=population_index, outcome_index=outcome_index)
    context = {
        'user_can_edit_attributes': user_subject.can_edit_attributes,
        'subject': subject,
        'publication': publication,
        'experiment': experiment,
        'experiment_population': experiment_population,
        'experiment_population_outcome': experiment_population_outcome,
        'experiment_population_outcome_form': experiment_population_outcome_form,
        'experiment_index': experiment_index,
        'population_index': population_index,
        'outcome_index': outcome_index,
        'data_formset': data_formset,
        'coordinates_formset': coordinates_formset,
        'date_formset': date_formset,
        'study_formset': study_formset,
        'EAV_formset': EAV_formset,
        'x_country_formset': x_country_formset,
        'path_to_shiny': get_path_to_shiny(request)
    }
    return render(request, 'publications/outcome.html', context)


def browse_by_intervention(request, subject, state, set='default', download='none'):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    if user.is_authenticated:
        if UserSubject.objects.filter(user=user, subject=subject).exists():
            subjects = subject.get_descendants(include_self=True)
        else:
            subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    else:
        subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    path_to_shiny = ''
    if (state == 'publications'):
        intervention = subject.intervention    # The root intervention for this subject (each subject can have its own classification of interventions)
        interventions = Intervention.objects.filter(pk=intervention.pk).get_descendants(include_self=True)
        if (set == 'default'):
            publications = Publication.objects.filter(subject__in=subjects)
            interventions = interventions.filter(experiment__publication__in=publications).get_ancestors(include_self=True)
    if (state == 'data'):
        path_to_shiny = get_path_to_shiny(request)
        data = Data.objects.filter(subject__in=subjects)
        interventions = Intervention.objects.filter(experiment__data__in=data).get_ancestors(include_self=True)
    interventions_count = interventions.count()
    # If the request is to download a CSV file with the number of publications per intervention
    if (download == 'CSV'):
        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="publications_per_intervention.csv"'
        # Write the CSV
        writer = csv.writer(response, quoting=csv.QUOTE_ALL)
        writer.writerow(['level_0', 'level_1', 'level_2', 'level_3', 'level_4', 'level_5', 'level_6', 'level_7', 'number_of_publications'])
        for intervention in interventions:
            level = intervention.level
            if intervention.code is not None and intervention.code != "":
                code = "{code} ".format(code = intervention.code)
            else:
                code = ""
            if (level == 0):
                level_0 = str(subject).capitalize()
            else:
                level_0 = ""
            if (level == 1):
                level_1 = "{code}{intervention}".format(code = code, intervention = intervention)
            else:
                level_1 = ""
            if (level == 2):
                level_2 = "{code}{intervention}".format(code = code, intervention = intervention)
            else:
                level_2 = ""
            if (level == 3):
                level_3 = "{code}{intervention}".format(code = code, intervention = intervention)
            else:
                level_3 = ""
            if (level == 4):
                level_4 = "{code}{intervention}".format(code = code, intervention = intervention)
            else:
                level_4 = ""
            if (level == 5):
                level_5 = "{code}{intervention}".format(code = code, intervention = intervention)
            else:
                level_5 = ""
            if (level == 6):
                level_6 = "{code}{intervention}".format(code = code, intervention = intervention)
            else:
                level_6 = ""
            if (level == 7):
                level_7 = "{code}{intervention}".format(code = code, intervention = intervention)
            else:
                level_7 = ""
            publications_count = Publication.objects.distinct().filter(subject__in=subjects, experiment__intervention__in=intervention.get_descendants(include_self=True)).values_list('pk', flat=True).count()
            writer.writerow([level_0, level_1, level_2, level_3, level_4, level_5, level_6, level_7, publications_count])
        return response
    context = {
        'subject': subject,  # Browse within this subject
        'interventions': interventions,
        'state': state,
        'set': set,
        'interventions_count': interventions_count,
        'path_to_shiny': path_to_shiny
    }
    if user.is_authenticated:
        if Publication.objects.filter(subject=subject).exists():
            status = get_status(user, subject)
            context.update(status)
    return render(request, 'publications/browse_by_intervention.html', context)


def browse_by_outcome(request, subject, state, set='default', download='none'):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    if user.is_authenticated:
        if UserSubject.objects.filter(user=user, subject=subject).exists():
            subjects = subject.get_descendants(include_self=True)
        else:
            subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    else:
        subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    path_to_shiny = ''
    if (state == 'publications'):
        outcome = subject.outcome  # The root outcome for this subject (each subject can have its own classification of outcomes)
        outcomes = Outcome.objects.get(pk=outcome.pk).get_descendants(include_self=True)
        if (set == 'default'):
            publications = Publication.objects.filter(subject__in=subjects)
            outcomes = outcomes.distinct().filter(
                Q(experimentpopulationoutcome__experiment_population__experiment__publication__in=publications) |
                Q(publicationpopulationoutcome__publication_population__publication__in=publications)
            ).get_ancestors(include_self=True)
    if (state == 'data'):
        path_to_shiny = get_path_to_shiny(request)
        data = Data.objects.filter(subject__in=subjects)
        outcomes = Outcome.objects.filter(experimentpopulationoutcome__data__in=data).get_ancestors(include_self=True)
    outcomes_count = outcomes.count()
    # If the request is to download a CSV file with the number of publications per outcome
    if (download == 'CSV'):
        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="publications_per_outcome.csv"'
        # Write the CSV
        writer = csv.writer(response, quoting=csv.QUOTE_ALL)
        writer.writerow(['level_0', 'level_1', 'level_2', 'level_3', 'level_4', 'level_5', 'level_6', 'level_7', 'number_of_publications'])
        for outcome in outcomes:
            level = outcome.level
            if outcome.code is not None and outcome.code != "":
                code = "{code} ".format(code = outcome.code)
            else:
                code = ""
            if (level == 0):
                level_0 = str(subject).capitalize()
            else:
                level_0 = ""
            if (level == 1):
                level_1 = "{code}{outcome}".format(code = code, outcome = outcome)
            else:
                level_1 = ""
            if (level == 2):
                level_2 = "{code}{outcome}".format(code = code, outcome = outcome)
            else:
                level_2 = ""
            if (level == 3):
                level_3 = "{code}{outcome}".format(code = code, outcome = outcome)
            else:
                level_3 = ""
            if (level == 4):
                level_4 = "{code}{outcome}".format(code = code, outcome = outcome)
            else:
                level_4 = ""
            if (level == 5):
                level_5 = "{code}{outcome}".format(code = code, outcome = outcome)
            else:
                level_5 = ""
            if (level == 6):
                level_6 = "{code}{outcome}".format(code = code, outcome = outcome)
            else:
                level_6 = ""
            if (level == 7):
                level_7 = "{code}{outcome}".format(code = code, outcome = outcome)
            else:
                level_7 = ""
            publications_count = Publication.objects.filter(
                subject__in=subjects
            ).filter(
                Q(experiment__experimentpopulation__experimentpopulationoutcome__outcome__in=outcome.get_descendants(include_self=True)) |
                Q(publicationpopulation__publicationpopulationoutcome__outcome__in=outcome.get_descendants(include_self=True))
            ).values_list('pk', flat=True).distinct().count()
            writer.writerow([level_0, level_1, level_2, level_3, level_4, level_5, level_6, level_7, publications_count])
        return response
    context = {
        'subject': subject,  # Browse within this subject
        'outcomes': outcomes,
        'state': state,
        'set': set,
        'outcomes_count': outcomes_count,
        'path_to_shiny': path_to_shiny
    }
    if user.is_authenticated:
        if Publication.objects.filter(subject=subject).exists():
            status = get_status(user, subject)
            context.update(status)
    return render(request, 'publications/browse_by_outcome.html', context)


def this_intervention(request, subject, state, intervention_pk, outcome_pk='default'):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    if user.is_authenticated:
        if UserSubject.objects.filter(user=user, subject=subject).exists():
            subjects = subject.get_descendants(include_self=True)
        else:
            subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    else:
        subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    path = ''
    path_to_shiny = ''
    this_intervention = Intervention.objects.get(pk=intervention_pk)
    interventions = Intervention.objects.filter(pk=intervention_pk).get_descendants(include_self=True)
    if outcome_pk != 'default':
        this_outcome = Outcome.objects.get(pk=outcome_pk)
        outcomes = Outcome.objects.filter(pk=outcome_pk).get_descendants(include_self=True)
    if state == 'publications':
        # Publications for this intervention (and its descendants)
        assessments = Assessment.objects.filter(subject__in=subjects, full_text_is_relevant=True)
        publications = Publication.objects.filter(subject__in=subjects, assessment__in=assessments)
        publications = publications.distinct().filter(experiment__intervention__in=interventions)
        # Filter these publications by outcome
        if outcome_pk != 'default':
            publications = publications.distinct().filter(
                    Q(experiment__experimentpopulation__experimentpopulationoutcome__outcome__in=outcomes) |
                    Q(publicationpopulation__publicationpopulationoutcome__outcome__in=outcomes)
                )
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by intervention and outcome.
            path = reverse('publications_x', args=(), kwargs={'subject': subject.slug, 'intervention_pk': this_intervention.pk, 'outcome_pk': this_outcome.pk})
        else:
            this_outcome = None
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by intervention (not outcome).
            path = reverse('publications_x', args=(), kwargs={'subject': subject.slug, 'intervention_pk': this_intervention.pk})

        # Outcomes for these publications
        # By publication (outcomes can be entered by publication or by experiment)
        publication_population_outcomes = PublicationPopulationOutcome.objects.filter(
            publication_population__publication__in=publications)
        # By experiment (outcomes can be entered by publication or by experiment)
        experiment_population_outcomes = ExperimentPopulationOutcome.objects.filter(
            experiment_population__experiment__publication__in=publications)
        # All outcomes for these publications (outcomes by publication OR outcomes by experiment)
        outcomes = Outcome.objects.distinct().filter(
                Q(experimentpopulationoutcome__in=experiment_population_outcomes) |
                Q(publicationpopulationoutcome__in=publication_population_outcomes)
            ).get_ancestors(include_self=True)

        # Countries for these publications
        countries = XCountry.objects.filter(publication_index__in=publications)

        # The number of publications by country
        q = countries.values_list('country__iso_alpha_3', 'publication_index').distinct()
        countries = list(chain(q))  # A list of tuples in the form [(country, publication)]. Chain is imported from itertools.
        countries = set(countries)  # Delete duplicate records, where a publication has the same country in both publication_country and experiment_country: set = unique tuples (and the list is now a dict)
        count_by_country = Counter(item[0] for item in countries)  # item[0] is country in (country, publication) and this counts the number of tuples for each country. Counter is imported from collections.
        count_by_country = json.dumps(count_by_country)  # Convert to JSON for use by JavaScript in the template
        count = publications.count()

    if state == 'data':
        path_to_shiny = get_path_to_shiny(request)

        # Data for this intervention (and its descendants)
        data = Data.objects.filter(subject__in=subjects)
        data = data.filter(experiment__intervention__in=interventions)
        # Filter these data by outcome
        if outcome_pk != 'default':
            data = data.filter(experiment_population_outcome__outcome__in=outcomes)
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by intervention and outcome.
            path = "{path_to_shiny}?subject={subject}&intervention={intervention}&outcome={outcome}&country=".format(path_to_shiny=path_to_shiny, subject=subject.pk, intervention=this_intervention.pk, outcome=this_outcome.pk)
        else:
            this_outcome = None
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by intervention (not outcome).
            path = "{path_to_shiny}?subject={subject}&intervention={intervention}&country=".format(path_to_shiny=path_to_shiny, subject=subject.pk, intervention=this_intervention.pk)

        # Outcomes for these data
        outcomes = Outcome.objects.distinct().filter(
                experimentpopulationoutcome__data__in=data
            ).get_ancestors(include_self=True)

        # Countries for these data
        experiments = Experiment.objects.filter(data__in=data)
        countries = XCountry.objects.filter(experiment_index__in=experiments)
        q = countries.values_list('country__iso_alpha_3', 'experiment_index__data').distinct()
        countries = list(chain(q))  # A list of tuples in the form [(country, publication)]. Chain is imported from itertools.
        countries = set(countries)  # Delete duplicate records, where a publication has the same country in both publication_country and experiment_country: set = unique tuples (and the list is now a dict)
        count_by_country = Counter(item[0] for item in countries)  # item[0] is country in (country, publication) and this counts the number of tuples for each country. Counter is imported from collections.
        count_by_country = json.dumps(count_by_country)  # Convert to JSON for use by JavaScript in the template
        count = data.count()

    context = {
        'subject': subject,
        'this_intervention': this_intervention,
        'this_outcome': this_outcome,
        'outcomes': outcomes,
        'count': count,
        'count_by_country': count_by_country,
        'path': path,
        'path_to_shiny': path_to_shiny,
        'state': state
    }
    if user.is_authenticated:
        if Publication.objects.filter(subject=subject).exists():
            status = get_status(user, subject)
            context.update(status)
    return render(request, 'publications/this_intervention.html', context)


def this_outcome(request, subject, state, outcome_pk, intervention_pk='default'):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    if user.is_authenticated:
        if UserSubject.objects.filter(user=user, subject=subject).exists():
            subjects = subject.get_descendants(include_self=True)
        else:
            subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    else:
        subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    path = ''
    path_to_shiny = ''
    this_outcome = Outcome.objects.get(pk=outcome_pk)
    outcomes = Outcome.objects.filter(pk=outcome_pk).get_descendants(include_self=True)
    if state == 'publications':
        # Records for these outcomes
        publication_population_outcomes = PublicationPopulationOutcome.objects.filter(outcome__in=outcomes)
        experiment_population_outcomes = ExperimentPopulationOutcome.objects.filter(outcome__in=outcomes)
        # Publications for these records
        assessments = Assessment.objects.filter(subject__in=subjects, full_text_is_relevant=True)
        publications = Publication.objects.distinct().filter(subject__in=subjects, assessment__in=assessments)
        publications = publications.distinct().filter(
                Q(experiment__experimentpopulation__experimentpopulationoutcome__in=experiment_population_outcomes) |
                Q(publicationpopulation__publicationpopulationoutcome__in=publication_population_outcomes)
            )
        # Filter these publications by intervention
        if intervention_pk != 'default':
            this_intervention = Intervention.objects.get(pk=intervention_pk)
            interventions = Intervention.objects.filter(pk=intervention_pk).get_descendants(include_self=True)
            # Publications for these interventions
            publications = publications.distinct().filter(experiment__intervention__in=interventions)
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by intervention and outcome.
            path = reverse('publications_x', args=(), kwargs={'subject': subject.slug, 'intervention_pk': this_intervention.pk, 'outcome_pk': this_outcome.pk})
        else:
            this_intervention = None
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by outcome (not intervention).
            path = reverse('publications_x', args=(), kwargs={'subject': subject.slug, 'outcome_pk': this_outcome.pk})
        # Interventions for these publications
        interventions = Intervention.objects.distinct().filter(
                experiment__publication__in=publications
            ).get_ancestors(include_self=True)
        # Countries for these publications
        countries = XCountry.objects.filter(publication_index__in=publications)
        # The number of publications by country
        q = countries.values_list('country__iso_alpha_3', 'publication_index').distinct()
        countries = list(chain(q))  # A list of tuples in the form [(country, publication)]. Chain is imported from itertools.
        countries = set(countries)  # Delete duplicate records, where a publication has the same country in both publication_country and experiment_country: set = unique tuples (and the list is now a dict)
        count_by_country = Counter(item[0] for item in countries)  # item[0] is country in (country, publication) and this counts the number of tuples for each country. Counter is imported from collections.
        count_by_country = json.dumps(count_by_country)  # Convert to JSON for use by JavaScript in the template
        count = publications.count()

    if state == 'data':
        path_to_shiny = get_path_to_shiny(request)

        # Data for this outcome (and its descendants)
        data = Data.objects.filter(subject__in=subjects)
        data = data.filter(experiment_population_outcome__outcome__in=outcomes)
        # Filter these data by intervention
        if intervention_pk != 'default':
            this_intervention = Intervention.objects.get(pk=intervention_pk)
            interventions = Intervention.objects.filter(pk=intervention_pk).get_descendants(include_self=True)
            data = data.filter(experiment__intervention__in=interventions)
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by intervention and outcome.
            path = "{path_to_shiny}?subject={subject}&intervention={intervention}&outcome={outcome}&country=".format(path_to_shiny=path_to_shiny, subject=subject.pk, intervention=this_intervention.pk, outcome=this_outcome.pk)
        else:
            this_intervention = None
            # Pass this path to the template for evidence-atlas.js to generate dynamic links by outcome (not intervention).
            path = "{path_to_shiny}?subject={subject}&outcome={outcome}&country=".format(path_to_shiny=path_to_shiny, subject=subject.pk, outcome=this_outcome.pk)
        # Interventions for these data
        interventions = Intervention.objects.distinct().filter(
                experiment__data__in=data
            ).get_ancestors(include_self=True)
        # Get the metadata that were entered at the lowest level (publication >
        # intervention > population > outcome) using Coalesce.
        data_countries = data.annotate(country_codes=Coalesce(
            'experiment_population_outcome__xcountry_outcome__country__iso_alpha_3',
            'experiment_population__xcountry_population__country__iso_alpha_3',
            'experiment__xcountry_experiment__country__iso_alpha_3',
            'publication__xcountry_publication__country__iso_alpha_3'
        )).filter()
        data_countries = data_countries.values_list('country_codes', 'pk')
        # Convert to a list of tuples and count the tuples for each country.
        data_countries = set(chain(data_countries))
        count_by_country = Counter(item[0] for item in data_countries)  # item[0] is country in (country, pk) and this counts the number of tuples for each country. Counter is imported from collections.
        count_by_country = json.dumps(count_by_country)  # Convert to JSON for use by JavaScript in the template
        count = data.count()

    context = {
        'subject': subject,
        'this_intervention': this_intervention,
        'this_outcome': this_outcome,
        'interventions': interventions,
        'count': count,
        'count_by_country': count_by_country,
        'path': path,
        'path_to_shiny': path_to_shiny,
        'state': state
    }
    if user.is_authenticated:
        if Publication.objects.filter(subject=subject).exists():
            status = get_status(user, subject)
            context.update(status)
    return render(request, 'publications/this_outcome.html', context)


def publications_x(request, subject, intervention_pk='default', outcome_pk='default', iso_a3='default'):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    if user.is_authenticated:
        if UserSubject.objects.filter(user=user, subject=subject).exists():
            subjects = subject.get_descendants(include_self=True)
        else:
            subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    else:
        subjects = subject.get_descendants(include_self=True).exclude(is_public=False)
    assessments = Assessment.objects.filter(subject__in=subjects, full_text_is_relevant=True)
    publications = Publication.objects.filter(subject__in=subjects, assessment__in=assessments)
    if iso_a3 != 'default':
        if iso_a3 != '-99':  # Disputed territories that are not our Countries model: Kosovo, Northern Cyprus, and Somaliland
            country = Country.objects.get(iso_alpha_3=iso_a3)
            publications = publications.distinct().filter(xcountry_publication_index__country=country)
    if outcome_pk != 'default':
        outcomes = Outcome.objects.filter(pk=outcome_pk).get_descendants(include_self=True)
        # Outcomes by experiment (outcomes can be entered by experiment or by publication)
        experiment_population_outcomes = ExperimentPopulationOutcome.objects.filter(outcome__in=outcomes)
        # Outcomes by publication (outcomes can be entered by experiment or by publication)
        publication_population_outcomes = PublicationPopulationOutcome.objects.filter(outcome__in=outcomes)
        # Publications for these outcomes
        publications = publications.distinct().filter(
                Q(experiment__experimentpopulation__experimentpopulationoutcome__in=experiment_population_outcomes) |
                Q(publicationpopulation__publicationpopulationoutcome__in=publication_population_outcomes)
            )
    if intervention_pk != 'default':
        interventions = Intervention.objects.filter(pk=intervention_pk).get_descendants(include_self=True)
        publications = publications.distinct().filter(experiment__intervention__in=interventions)
    publications = publications.order_by('title')
    if iso_a3 == '-99':  # Disputed territories that are not our Countries model: Kosovo, Northern Cyprus, and Somaliland
        publications = Publication.objects.none()
    page = request.GET.get('page', 1)
    paginator = Paginator(publications, 10)
    try:
        publications = paginator.page(page)
    except PageNotAnInteger:
        publications = paginator.page(1)
    except EmptyPage:
        publications = paginator.page(paginator.num_pages)
    context = {
        'subject': subject,
        'publications': publications
    }
    if user.is_authenticated:
        if Publication.objects.filter(subject=subject).exists():
            status = get_status(user, subject)
            context.update(status)
    return render(request, 'publications/publications.html', context)


def kappa(request, subject):
    user = request.user
    data = request.POST or None
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    user_subjects = UserSubject.objects.filter(subject=subject)
    kappa = None
    kappa_form = KappaForm(data=data)
    kappa_form.fields['user_1'] = ModelChoiceField(queryset=User.objects.filter(usersubject__in=user_subjects))
    kappa_form.fields['user_2'] = ModelChoiceField(queryset=User.objects.filter(usersubject__in=user_subjects))
    intervention = subject.intervention
    intervention = Intervention.objects.get(pk=intervention.pk)
    kappa_form.fields['intervention'] = TreeNodeChoiceField(required=False, queryset=intervention.get_descendants(include_self=True), level_indicator="---", initial=intervention.pk)
    outcome = subject.outcome
    outcome = Outcome.objects.get(pk=outcome.pk)
    kappa_form.fields['outcome'] = TreeNodeChoiceField(required=False, queryset=outcome.get_descendants(include_self=True), level_indicator="---", initial=outcome.pk)
    if request.method == 'POST':
        form = kappa_form
        if form.is_valid():
            user_1 = form.cleaned_data.get('user_1')
            user_2 = form.cleaned_data.get('user_2')
            users = User.objects.filter(pk=user_1.pk) | User.objects.filter(pk=user_2.pk)  # The "|" operator requires querysets, not objects, so we use "filter" here and "get" below (we cannot use "get" before we use "|".)
            user_1 = User.objects.get(pk=user_1.pk)
            user_2 = User.objects.get(pk=user_2.pk)
            percent = form.cleaned_data.get('percent')
            number = form.cleaned_data.get('number')
            filter_1 = request.POST.get('filter_1')
            filter_2 = request.POST.get('filter_2')
            intervention = form.cleaned_data.get('intervention')
            outcome = form.cleaned_data.get('outcome')

            # Kappa analysis at Stage 1 (titles/abstracts) or Stage 2 (full texts)
            if 'stage_1' in request.POST or 'stage_2' in request.POST:

                # Kappa analysis at Stage 1 (titles/abstracts)
                if 'stage_1' in request.POST:
                    stage = "titles/abstracts"
                    total_n = Publication.objects.filter(subject=subject).count()
                    # Publications that user_1 assessed
                    user_1_publications = Publication.objects.filter(
                        subject=subject,
                        assessment__in=Assessment.objects.filter(user=user_1)
                    )
                    # Publications that user_2 assessed
                    user_2_publications = Publication.objects.filter(
                        subject=subject,
                        assessment__in=Assessment.objects.filter(user=user_2)
                    )
                    # Publications that both users assessed
                    publications = user_1_publications & user_2_publications
                    n = publications.count()
                    # Get a subset of these publications (based on form input).
                    if filter_1 == "filter_by_number":
                        target_n = number
                    else:  # If filter_2 == "filter_by_percent":
                        target_proportion = percent / 100
                        target_n = round(total_n * target_proportion)
                    if filter_2 == "least_recent":
                        assessments = Assessment.objects.filter(user=user_1, publication__in=publications).order_by('updated')
                    else:  # if filter_2 == "most_recent":
                        assessments = Assessment.objects.filter(user=user_1, publication__in=publications).order_by('-updated')
                    target_assessments = list(assessments[:target_n])
                    publications = Publication.objects.filter(assessment__in=target_assessments)
                    n = publications.count()
                    percent = (n / total_n) * 100
                    # Publications that user_1 included
                    user_1_included = publications.filter(
                        assessment__in=Assessment.objects.filter(user=user_1, is_relevant=True)
                    )
                    # Publications that user_2 included
                    user_2_included = publications.filter(
                        assessment__in=Assessment.objects.filter(user=user_2, is_relevant=True)
                    )
                    # Publications that user_1 excluded
                    user_1_excluded = publications.filter(
                        assessment__in=Assessment.objects.filter(user=user_1, is_relevant=False)
                    )
                    # Publications that user_2 excluded
                    user_2_excluded = publications.filter(
                        assessment__in=Assessment.objects.filter(user=user_2, is_relevant=False)
                    )

                # Kappa analysis at Stage 2 (full texts)
                elif 'stage_2' in request.POST:
                    stage = "full texts"
                    # Publications that any user assessed as relevant at Stage 1
                    total_publications = Publication.objects.filter(
                        subject = subject,
                        assessment__in=Assessment.objects.filter(is_relevant=True)
                    ).distinct()
                    total_n = total_publications.count()
                    # Publications that user_1 assessed at Stage 2
                    user_1_publications = Publication.objects.filter(
                        subject=subject,
                        assessment__in=Assessment.objects.filter(user=user_1, full_text_is_relevant__isnull=False)
                    )
                    # Publications that user_2 assessed at Stage 2
                    user_2_publications = Publication.objects.filter(
                        subject=subject,
                        assessment__in=Assessment.objects.filter(user=user_2, full_text_is_relevant__isnull=False)
                    )
                    # Publications that both users assessed at Stage 2
                    publications = user_1_publications & user_2_publications
                    n = publications.count()
                    # Get a subset of these publications (based on form input).
                    if filter_1 == "filter_by_number":
                        target_n = number
                    else:  # If filter_2 == "filter_by_percent":
                        target_proportion = percent / 100
                        target_n = round(total_n * target_proportion)
                    if filter_2 == "least_recent":
                        assessments = Assessment.objects.filter(user=user_1, publication__in=publications).order_by('updated')
                    else:  # if filter_2 == "most_recent":
                        assessments = Assessment.objects.filter(user=user_1, publication__in=publications).order_by('-updated')
                    target_assessments = list(assessments[:target_n])
                    publications = Publication.objects.filter(assessment__in=target_assessments)
                    n = publications.count()
                    percent = (n / total_n) * 100
                    # Publications that user_1 included
                    try:
                        user_1_included = publications.filter(assessment__in=Assessment.objects.filter(user=user_1, full_text_is_relevant=True))
                    except:
                        user_1_included = None
                    # Publications that user_2 included
                    try:
                        user_2_included = publications.filter(assessment__in=Assessment.objects.filter(user=user_2, full_text_is_relevant=True))
                    except:
                        user_2_included = None
                    # Publications that user_1 excluded
                    try:
                        user_1_excluded = publications.filter(assessment__in=Assessment.objects.filter(user=user_1, full_text_is_relevant=False))
                    except:
                        user_1_excluded = None
                    # Publications that user_2 excluded
                    try:
                        user_2_excluded = publications.filter(assessment__in=Assessment.objects.filter(user=user_2, full_text_is_relevant=False))
                    except:
                        user_2_excluded = None

                # Publications that both users included
                both_included = user_1_included & user_2_included
                # Publications that both users excluded
                both_excluded = user_1_excluded & user_2_excluded
                # Publications that user_1 included but user_2_excluded
                try:
                    only_user_1_included = user_1_included.exclude(pk__in=user_2_included)
                except:
                    only_user_1_included = None
                # Publications that user_2 included but user_1_excluded
                try:
                    only_user_2_included = user_2_included.exclude(pk__in=user_1_included)
                except:
                    only_user_2_included = None

                # Data for Kappa analysis
                try:
                    a = both_included.count()
                except:
                    a = 0
                try:
                    b = only_user_2_included.count()
                except:
                    b = 0
                try:
                    c = only_user_1_included.count()
                except:
                    c = 0
                try:
                    d = both_excluded.count()
                except:
                    d = 0

            # Kappa analysis at Stage 3a (interventions) or Stage 3b (outcomes)
            elif ('stage_3a' in request.POST or 'stage_3b' in request.POST):
                # Publications that any user included at Stage 2
                total_publications = Publication.objects.filter(
                    subject = subject,
                    assessment__in=Assessment.objects.filter(full_text_is_relevant=True)
                ).distinct()
                total_n = total_publications.count()
                # Publications that user_1 included at Stage 2
                user_1_publications = Publication.objects.filter(
                    subject=subject,
                    assessment__in=Assessment.objects.filter(user=user_1, full_text_is_relevant=True)
                )
                # Publications that user_2 included at Stage 2
                user_2_publications = Publication.objects.filter(
                    subject=subject,
                    assessment__in=Assessment.objects.filter(user=user_2, full_text_is_relevant=True)
                )
                # Publications that both users included at Stage 2
                publications = user_1_publications & user_2_publications
                n = publications.count()
                # Get a subset of these publications (based on form input).
                if filter_1 == "filter_by_number":
                    target_n = number
                else:  # If filter_2 == "filter_by_percent":
                    target_proportion = percent / 100
                    target_n = round(total_n * target_proportion)
                if filter_2 == "least_recent":
                    assessments = Assessment.objects.filter(user=user_1, publication__in=publications).order_by('updated')
                else:  # if filter_2 == "most_recent":
                    assessments = Assessment.objects.filter(user=user_1, publication__in=publications).order_by('-updated')
                target_assessments = list(assessments[:target_n])
                publications = Publication.objects.filter(assessment__in=target_assessments)
                n = publications.count()
                percent = (n / total_n) * 100
                a = 0
                b = 0
                c = 0
                d = 0
                only_user_1_included = []
                only_user_2_included = []
                show_disagreements = False

                # Kappa analysis at Stage 3a (interventions)
                if 'stage_3a' in request.POST:
                    stage = "interventions"
                    intervention = Intervention.objects.get(pk=intervention.pk)
                    if intervention.is_leaf_node():
                        show_disagreements = True
                    interventions = intervention.get_descendants(include_self=True)
                    interventions = interventions.filter(experiment__publication__in=publications).distinct()
                    n = n * interventions.count()
                    # Did both users select the same interventions for the same publications?
                    for intervention in interventions:
                        # Publications for which neither user selected this intervention
                        publications_without = publications.exclude(experiment__intervention=intervention).distinct()
                        d = d + publications_without.count()
                        # Publications for which either or both selected this intervention
                        publications_with = publications.filter(experiment__intervention=intervention).distinct()
                        for publication in publications_with:
                            user_1_included = Experiment.objects.filter(intervention=intervention, publication=publication, user=user_1).exists()
                            user_2_included = Experiment.objects.filter(intervention=intervention, publication=publication, user=user_2).exists()
                            if user_1_included and user_2_included:
                                a = a + 1
                            if user_1_included and not user_2_included:
                                c = c + 1
                                only_user_1_included.append(publication.pk)
                            if user_2_included and not user_1_included:
                                b = b + 1
                                only_user_2_included.append(publication.pk)
                            if not user_1_included and not user_2_included:
                                d = d + 1

                # Kappa analysis at Stage 3b (outcomes)
                if 'stage_3b' in request.POST:
                    stage = "outcomes"
                    outcome = Outcome.objects.get(pk=outcome.pk)
                    if outcome.is_leaf_node():
                        show_disagreements = True
                    outcomes = outcome.get_descendants(include_self=True)
                    outcomes = outcomes.filter(
                        Q(experimentpopulationoutcome__experiment_population__experiment__publication__in=publications) |
                        Q(publicationpopulationoutcome__publication_population__publication__in=publications)
                    ).distinct()
                    n = n * outcomes.count()
                    # Did both users select the same interventions for the same publications?
                    for outcome in outcomes:
                        # Publications for which either or both selected this intervention
                        publications_with = publications.filter(
                            Q(experiment__experimentpopulation__experimentpopulationoutcome__outcome=outcome) |
                            Q(publicationpopulation__publicationpopulationoutcome__outcome=outcome)
                        ).distinct()
                        # Publications for which neither user selected this intervention
                        publications_without = publications.exclude(id__in=publications_with)
                        d = d + publications_without.count()
                        for publication in publications_with:
                            user_1_included_in_experiment = ExperimentPopulationOutcome.objects.filter(outcome=outcome, experiment_population__experiment__user=user_1, experiment_population__experiment__publication=publication).exists()
                            user_1_included_in_publication = PublicationPopulationOutcome.objects.filter(outcome=outcome, user=user_1, publication_population__publication=publication).exists()
                            user_1_included = user_1_included_in_experiment or user_1_included_in_publication
                            user_2_included_in_experiment = ExperimentPopulationOutcome.objects.filter(outcome=outcome, experiment_population__experiment__user=user_2, experiment_population__experiment__publication=publication).exists()
                            user_2_included_in_publication = PublicationPopulationOutcome.objects.filter(outcome=outcome, user=user_2, publication_population__publication=publication).exists()
                            user_2_included = user_2_included_in_experiment or user_2_included_in_publication
                            if user_1_included and user_2_included:
                                a = a + 1
                            if user_1_included and not user_2_included:
                                c = c + 1
                                only_user_1_included.append(publication.pk)
                            if user_2_included and not user_1_included:
                                b = b + 1
                                only_user_2_included.append(publication.pk)
                            if not user_1_included and not user_2_included:
                                d = d + 1

                if show_disagreements:
                    if not only_user_1_included:
                        only_user_1_included = None
                    else:
                        only_user_1_included = publications.filter(pk__in=list(set(only_user_1_included)))
                    if not only_user_2_included:
                        only_user_2_included = None
                    else:
                        only_user_2_included = publications.filter(pk__in=list(set(only_user_2_included)))
                else:
                    only_user_1_included = None
                    only_user_2_included = None

            # Kappa analysis
            if n > 0:
                rm1 = a + b
                rm2 = c + d
                cm1 = a + c
                cm2 = b + d
                agreement = (a + d) / n
                expected_agreement = ( ((cm1 * rm1) / n) + ((cm2 * rm2) / n) ) / n
                if expected_agreement < 1:
                    kappa = (agreement - expected_agreement) / (1 - expected_agreement)
                    if kappa > 1:
                        kappa = 1.0
                else:
                    kappa = 1.0

    context = {
        'subject': subject,
        'user_subject': user_subject,
        'form': kappa_form
    }
    if kappa is not None:
        context.update({
            'stage': stage,
            'a': a,
            'b': b,
            'c': c,
            'd': d,
            'n': n,
            'total_n': total_n,
            'only_user_1_included': only_user_1_included,
            'only_user_2_included': only_user_2_included,
            'kappa': round(kappa, 2),
            'percent_agreement': round(agreement, 2),
            'percent': round(percent, 2)
        })
    return render(request, 'publications/kappa.html', context)


def get_status(user, subject):
    """
    Publications should be assessed in a random order, but each user should see
    the same order from session to session. Therefore, a random assessment_order
    is created for each user (for each subject), and it is saved in the database.
    """

    if UserSubject.objects.filter(user=user, subject=subject).exists():  # If the user has permission to work on this subject

        # If an assessment_order has been created for this user and subject, get it from the database.
        if AssessmentStatus.objects.filter(user=user, subject=subject).exists():
            item = AssessmentStatus.objects.get(user=user, subject=subject)
            assessment_order = literal_eval(item.assessment_order)
            next_assessment = item.next_assessment
            completed_assessments = literal_eval(item.completed_assessments)

            # If new publications have been added to the database, then randomly append their pks to the end of assessment_order.
            max_assessment_pks = max(assessment_order)
            pks = Publication.objects.filter(subject=subject).values_list('pk', flat=True)
            max_pks = max(pks)

            if max_assessment_pks < max_pks:
                new_publications = list(pks)
                new_publications = list(set(new_publications) - set(assessment_order))
                shuffle(new_publications)
                assessment_order = assessment_order + new_publications
                item.assessment_order = assessment_order
                item.save()

            #TODO: If old publications have been deleted from the database (this should not happen in production), then delete their pks from assessment_order.

        # If an assessment_order has not been created for this user and subject, create it and save it in the database.
        else:
            pks = Publication.objects.filter(subject=subject).values_list('pk', flat=True)
            assessment_order = list(pks)
            shuffle(assessment_order)
            next_assessment = assessment_order[0]
            previous_full_text_assessment = -1
            completed_assessments = []
            item = AssessmentStatus(
                subject=subject,
                user=user,
                assessment_order=assessment_order,
                next_assessment=next_assessment,
                previous_full_text_assessment=previous_full_text_assessment,
                completed_assessments=completed_assessments
            )
            item.save()
        item = AssessmentStatus.objects.get(user=user, subject=subject)
        publications_count = len(assessment_order)
        publications_assessed_count = len(completed_assessments)
        if publications_count != 0:
            publications_assessed_percent = int(publications_assessed_count / publications_count * 100)
        else:
            publications_assessed_percent = 100

        # Count the publications that have been included at title/abstract stage ("full texts") by this user.
        if Publication.objects.filter(
            assessment__in=Assessment.objects.filter(
                subject=subject, user=user, is_relevant=True
            )
        ).exists():
            full_texts = Publication.objects.filter(
                assessment__in=Assessment.objects.filter(
                    subject=subject, user=user, is_relevant=True
                )
            )
            full_texts_count = full_texts.values_list('pk', flat=True).count()
        else:
            full_texts_count = 0

        # Count the publications that have been included at title/abstract stage ("full texts") by any user.
        if (Publication.objects.filter(
            subject=subject,
            assessment__in=Assessment.objects.filter(is_relevant=True)
        ).distinct().exists()):
            all_full_texts = Publication.objects.filter(
                subject=subject,
                assessment__in=Assessment.objects.filter(is_relevant=True)
            ).distinct()
            all_full_texts_count = all_full_texts.values_list('pk', flat=True).count()
        else:
            all_full_texts_count = 0

        # Count the publications that have been included or excluded at full-text stage, but not necessarily marked as completed), by this user.
        if Publication.objects.filter(
            assessment__in=Assessment.objects.filter(
                subject=subject, user=user, full_text_is_relevant__isnull=False
            )
        ).exists():
            full_texts_assessed = Publication.objects.filter(
                assessment__in=Assessment.objects.filter(
                    subject=subject, user=user, full_text_is_relevant__isnull=False
                )
            )
            full_texts_assessed_count = full_texts_assessed.values_list('pk', flat=True).count()
        else:
            full_texts_assessed_count = 0
        if full_texts_count != 0:
            full_texts_assessed_percent = int(full_texts_assessed_count / full_texts_count * 100)
        else:
            full_texts_assessed_percent = 100

        # Count the publications that have been included or excluded at full-text stage, but not necessarily marked as completed), by any user.
        if Publication.objects.filter(
            assessment__in=Assessment.objects.filter(
                subject=subject, full_text_is_relevant__isnull=False
            )
        ).exists():
            all_full_texts_assessed = Publication.objects.filter(
                assessment__in=Assessment.objects.filter(
                    subject=subject, full_text_is_relevant__isnull=False
                )
            ).distinct()
            all_full_texts_assessed_count = all_full_texts_assessed.values_list('pk', flat=True).count()
        else:
            all_full_texts_assessed_count = 0
        if all_full_texts_count != 0:
            all_full_texts_assessed_percent = int(all_full_texts_assessed_count / all_full_texts_count * 100)
        else:
            all_full_texts_assessed_percent = 100

        # Count the publications that have been marked as completed at full-text stage by this user.
        if Publication.objects.filter(
            assessment__in=Assessment.objects.filter(
                subject=subject, user=user, is_completed=True
            )
        ).exists():
            full_texts_completed = Publication.objects.filter(
                assessment__in=Assessment.objects.filter(
                    subject=subject, user=user, is_completed=True
                )
            )
            full_texts_completed_count = full_texts_completed.values_list('pk', flat=True).count()
        else:
            full_texts_completed_count = 0
        if full_texts_count != 0:
            full_texts_completed_percent = int(full_texts_completed_count / full_texts_count * 100)
        else:
            full_texts_completed_percent = 100

        # Count the publications that have been marked as completed at full-text stage by any user.
        if Publication.objects.filter(
            assessment__in=Assessment.objects.filter(
                subject=subject, is_completed=True
            )
        ).exists():
            all_full_texts_completed = Publication.objects.filter(
                assessment__in=Assessment.objects.filter(
                    subject=subject, is_completed=True
                )
            ).distinct()
            all_full_texts_completed_count = all_full_texts_completed.values_list('pk', flat=True).count()
        else:
            all_full_texts_completed_count = 0
        if all_full_texts_count != 0:
            all_full_texts_completed_percent = int(all_full_texts_completed_count / all_full_texts_count * 100)
        else:
            all_full_texts_completed_percent = 100

        # Count the publications that have been included or excluded by this user and the user_for_comparison (for Kappa analysis).
        abstracts_assessed_by_user_2_count = None
        abstracts_assessed_by_both_users_count = None
        abstracts_assessed_by_both_users_percent = None
        abstracts_assessed_by_both_users_percent_of_user_2 = None
        full_texts_assessed_by_user_2_count = None
        full_texts_assessed_by_both_users_count = None
        full_texts_assessed_by_both_users_percent = None
        full_texts_assessed_by_both_users_percent_of_user_2 = None
        user_subject = UserSubject.objects.get(user=user, subject=subject)
        if user_subject.user_for_comparison:
            user_1 = user
            user_2 = user_subject.user_for_comparison

            # Titles/abstracts (Stage 1)
            # Publications that user_1 assessed at Stage 2
            user_1_publications = Publication.objects.filter(
                subject=subject,
                assessment__in=Assessment.objects.filter(user=user_1)
            )
            # Publications that user_2 assessed at Stage 2
            user_2_publications = Publication.objects.filter(
                subject=subject,
                assessment__in=Assessment.objects.filter(user=user_2)
            )
            abstracts_assessed_by_user_2_count = user_2_publications.values_list('pk', flat=True).count()
            # Publications that both users assessed at Stage 1
            abstracts_assessed_by_both_users = user_1_publications & user_2_publications
            abstracts_assessed_by_both_users_count = abstracts_assessed_by_both_users.values_list('pk', flat=True).count()
            if publications_count != 0:
                abstracts_assessed_by_both_users_percent = round(abstracts_assessed_by_both_users_count / publications_count * 100)
            else:
                abstracts_assessed_by_both_users_percent = 100
            if abstracts_assessed_by_user_2_count != 0:
                abstracts_assessed_by_both_users_percent_of_user_2 = round(abstracts_assessed_by_both_users_count / abstracts_assessed_by_user_2_count * 100)
            else:
                abstracts_assessed_by_both_users_percent_of_user_2 = 100

            # Full texts (Stage 2)
            # Publications that user_1 assessed at Stage 2
            user_1_publications = Publication.objects.filter(
                subject=subject,
                assessment__in=Assessment.objects.filter(
                    user=user_1, full_text_is_relevant__isnull=False
                )
            )
            # Publications that user_2 assessed at Stage 2
            user_2_publications = Publication.objects.filter(
                subject=subject,
                assessment__in=Assessment.objects.filter(
                    user=user_2, full_text_is_relevant__isnull=False
                )
            )
            full_texts_assessed_by_user_2_count = user_2_publications.values_list('pk', flat=True).count()
            # Publications that both users assessed at Stage 2
            full_texts_assessed_by_both_users = user_1_publications & user_2_publications
            full_texts_assessed_by_both_users_count = full_texts_assessed_by_both_users.values_list('pk', flat=True).count()
            if all_full_texts_count != 0:
                full_texts_assessed_by_both_users_percent = round(full_texts_assessed_by_both_users_count / all_full_texts_count * 100)
            else:
                full_texts_assessed_by_both_users_percent = 100
            if full_texts_assessed_by_user_2_count != 0:
                full_texts_assessed_by_both_users_percent_of_user_2 = round(full_texts_assessed_by_both_users_count / full_texts_assessed_by_user_2_count * 100)
            else:
                full_texts_assessed_by_both_users_percent_of_user_2 = 100
        status = {
            'user_subject': UserSubject.objects.filter(user=user, subject=subject).exists(),
            'item': item,
            'publications_count': publications_count,
            'publications_assessed_count': publications_assessed_count,
            'publications_assessed_percent': publications_assessed_percent,
            'abstracts_assessed_by_user_2_count': abstracts_assessed_by_user_2_count,
            'abstracts_assessed_by_both_users_count': abstracts_assessed_by_both_users_count,
            'abstracts_assessed_by_both_users_percent': abstracts_assessed_by_both_users_percent,
            'abstracts_assessed_by_both_users_percent_of_user_2': abstracts_assessed_by_both_users_percent_of_user_2,
            'full_texts_count': full_texts_count,  # "relevant publications" (at title/abstract stage) = "full texts" (publications that need to be screened at full-text stage)
            'full_texts_assessed_count': full_texts_assessed_count,
            'full_texts_assessed_percent': full_texts_assessed_percent,
            'full_texts_completed_count': full_texts_completed_count,
            'full_texts_completed_percent': full_texts_completed_percent,
            'all_full_texts_assessed_count': all_full_texts_assessed_count,
            'all_full_texts_assessed_percent': all_full_texts_assessed_percent,
            'all_full_texts_completed_count': all_full_texts_completed_count,
            'all_full_texts_completed_percent': all_full_texts_completed_percent,
            'all_full_texts_count': all_full_texts_count,
            'full_texts_assessed_by_user_2_count': full_texts_assessed_by_user_2_count,
            'full_texts_assessed_by_both_users_count': full_texts_assessed_by_both_users_count,
            'full_texts_assessed_by_both_users_percent': full_texts_assessed_by_both_users_percent,
            'full_texts_assessed_by_both_users_percent_of_user_2': full_texts_assessed_by_both_users_percent_of_user_2,
            'next_assessment': next_assessment
        }
        return(status)
    else:  # If the user does not have permission to work on this subject
        return()


# This method is needed (as are the lists of completed_assessments), because
# we need to maintain a unique random order for each user. Otherwise, we could
# use querysets for navigation, as we do for full_text_navigation.
def get_next_assessment(publication_pk, next_pk, assessment_order, completed_assessments):
    next_assessment = next_pk
    uncompleted_assessments = list(set(assessment_order) - set(completed_assessments))
    if next_assessment in uncompleted_assessments:
        return(next_assessment)
    else:
        if uncompleted_assessments:
            next_assessment = uncompleted_assessments[0]
            if next_assessment == publication_pk:  # Allow users to skip this publication.
                try:
                    next_assessment = uncompleted_assessments[uncompleted_assessments.index(next_assessment) + 1]
                except:
                    next_assessment = uncompleted_assessments[0]
        return(next_assessment)


def get_path_to_shiny(request):
    current_site = get_current_site(request)
    domain = current_site.domain
    if domain == "dev.metadataset.com":
        path_to_shiny = "http://metadataset.shinyapps.io/meta-analysis/"
        #path_to_shiny = "http://shiny.metadataset.com/meta-analysis/"
    else:
        path_to_shiny = "http://metadataset.shinyapps.io/meta-analysis/"
    return(path_to_shiny)


@login_required
def full_text_navigation(request, subject, direction, state, users, publication_pk='default'):
    user = request.user
    subject = Subject.objects.get(slug=subject)
    user_subject = get_object_or_404(UserSubject, user=user, subject=subject)  # Check if this user has permission to work on this subject.
    user_subjects = UserSubject.objects.filter(subject=subject)
    if (users=='this_user'):
        users = User.objects.filter(pk=user.pk)
    else:
        users = User.objects.filter(usersubject__subject=subject)
    # Get the publication from which to calculate "next" and "previous" (either the previous_full_text_assessment or the publication from which the request was sent).
    if (publication_pk=='default'):
        previous_full_text_assessment = AssessmentStatus.objects.get(
            subject=subject, user=user).previous_full_text_assessment
        if (previous_full_text_assessment == -1):  # If this is a new user, the initial value will be -1.
            previous_full_text_assessment = Publication.objects.filter(
                subject=subject
            ).order_by('-title').values_list('pk', flat=True)[0]
            assessment_status = AssessmentStatus.objects.get(
                subject=subject, user=user)
            assessment_status.previous_full_text_assessment = previous_full_text_assessment
            assessment_status.save
        publication = Publication.objects.get(pk=previous_full_text_assessment)
    else:
        publication = Publication.objects.get(pk=int(publication_pk))
    # Get all the publications.
    publications = Publication.objects.filter(subject=subject)
    # Exclude this publication.
    publications = publications.exclude(pk=publication.pk)
    if (state == 'all'):
        # Publications for this subject that were included at title/abstract stage (is_relevant=True)
        if publications.filter(assessment__in=Assessment.objects.filter(
            subject=subject, user__in=users, is_relevant=True
        )).exists():
            publications = publications.filter(assessment__in=Assessment.objects.filter(
                subject=subject, user__in=users, is_relevant=True
            ))
    elif (state == 'not_completed'):
        # Publications for this subject that were included at title/abstract stage but have not yet been marked as completed
        if publications.filter(assessment__in=Assessment.objects.filter(
            subject=subject, user__in=users, is_relevant=True, is_completed=False
        )).exists():
            publications = publications.filter(assessment__in=Assessment.objects.filter(
                subject=subject, user__in=users, is_relevant=True, is_completed=False
            ))
            if (users.count() > 1):
                # Exclude publications that have been assessed by at least one person (even if they have not been assessed by other people).
                if publications.exclude(assessment__in=Assessment.objects.filter(
                    subject=subject, user__in=users, is_completed=True
                )).exists():
                    publications = publications.exclude(assessment__in=Assessment.objects.filter(
                        subject=subject, user__in=users, is_completed=True
                    ))
    elif (state == 'not_assessed'):
        # Publications for this subject that were included at title/abstract stage but have not yet been assigned an intervention or excluded at full-text stage
        if publications.filter(assessment__in=Assessment.objects.filter(
            subject=subject, user__in=users, is_relevant=True, full_text_is_relevant=None
        )).exists():
            publications = publications.filter(assessment__in=Assessment.objects.filter(
                subject=subject, user__in=users, is_relevant=True, full_text_is_relevant=None
            ))
            if (users.count() > 1):
                # Exclude publications that have been assessed by at least one person (even if they have not been assessed by other people).
                if publications.exclude(assessment__in=Assessment.objects.filter(
                    subject=subject, user__in=users
                ).filter(
                    Q(full_text_is_relevant=True) | Q(full_text_is_relevant=False)
                )).exists():
                    publications = publications.exclude(assessment__in=Assessment.objects.filter(
                        subject=subject, user__in=users
                    ).filter(
                        Q(full_text_is_relevant=True) | Q(full_text_is_relevant=False)
                    ))
    elif (state == 'kappa'):
        # If screening full texts for Kappa analysis (publications that the user_for_comparison has already assessed at full-text stage)
        user_for_comparison = user_subject.user_for_comparison
        if publications.filter(assessment__in=Assessment.objects.filter(
            subject=subject, user=user_for_comparison, full_text_is_relevant__isnull=False
        )).exists():
            publications = publications.filter(assessment__in=Assessment.objects.filter(
                subject=subject, user=user_for_comparison, full_text_is_relevant__isnull=False
            ))
    # Go to the next or previous publication, in alphabetical order.
    if (direction == 'next'):
        try:
            publication_pk = publications.filter(
                title__gte=publication.title  # Titles later in the alphabet
            ).order_by('title').values_list('pk', flat=True)[0]
        except:
            publication_pk = publications.order_by('title').values_list('pk', flat=True)[0]
    elif (direction == 'previous'):
        try:
            publication_pk = publications.filter(
                title__lte=publication.title  # Titles earlier in the alphabet
            ).order_by('-title').values_list('pk', flat=True)[0]
        except:
            publication_pk = publications.order_by('-title').values_list('pk', flat=True)[0]
    # Update the current full_text_assessment
    assessment_status = AssessmentStatus.objects.get(subject=subject, user=user)
    assessment_status.previous_full_text_assessment = publication_pk
    assessment_status.save()
    return redirect('publication', subject=subject.slug, publication_pk=publication_pk)
